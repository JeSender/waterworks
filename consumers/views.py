from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from .decorators import (
    get_client_ip, get_user_agent, is_admin_user, is_superuser_only,
    consumer_edit_permission_required, disconnect_permission_required,
    user_management_permission_required, system_settings_permission_required,
    billing_permission_required, reports_permission_required, view_only_for_admin,
    rate_limit_login
)
from django.db.models import Q, Max, Count, Sum, OuterRef, Subquery, Value
from django.db.models.functions import Concat, TruncMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date
try:
    from dateutil.relativedelta import relativedelta
except Exception:
    # Fallback: approximate relativedelta by using a timedelta of ~30 days per month
    # This keeps existing subtraction usages like relativedelta(months=5) working
    from datetime import timedelta as _td
    def relativedelta(months=0, **kwargs):
        return _td(days=30 * int(months))
from decimal import Decimal, InvalidOperation
import uuid
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from .models import (
    Consumer, Barangay, Purok, MeterReading, Bill, SystemSetting, Payment,
    StaffProfile, UserLoginEvent, MeterBrand, PasswordResetToken, UserActivity
)
from .forms import ConsumerForm


# Helper function to get previous confirmed reading
def get_previous_reading(consumer):
    """Get the most recent confirmed meter reading for a consumer."""
    latest_reading = MeterReading.objects.filter(
        consumer=consumer,
        is_confirmed=True
    ).order_by('-reading_date', '-created_at').first()

    return latest_reading.reading_value if latest_reading else 0


# Helper function to calculate water bill
def calculate_water_bill(consumer, consumption):
    """
    Calculate water bill using TIERED rate structure from System Settings.

    Returns: (average_rate, total_amount, breakdown)
    - average_rate: Effective rate per cubic meter
    - total_amount: Total bill amount
    - breakdown: Dict with tier-by-tier calculation details
    """
    from .utils import calculate_tiered_water_bill

    # Use tiered calculation from utils
    total_amount, average_rate, breakdown = calculate_tiered_water_bill(
        consumption=consumption,
        usage_type=consumer.usage_type
    )

    return float(average_rate), float(total_amount), breakdown


# ============================================================================
# API VIEW: MOBILE APP METER READING SUBMISSION
# ============================================================================
# FLOW OVERVIEW:
# 1. Mobile app (field staff) submits meter reading anytime
# 2. Reading is saved with is_confirmed=TRUE (auto-confirmed)
# 3. Bill is automatically generated with status='Pending'
# 4. Bill appears in Inquire/Payment page for payment processing
#
# NOTE: Manual confirmation has been removed as it was redundant.
# Bills are now generated immediately upon meter reading submission.
# ============================================================================
@csrf_exempt  # Be careful with CSRF in production, consider using proper tokens for mobile apps
def api_submit_reading(request):
    """
    API endpoint for Android app to submit meter readings.

    IMPORTANT: Readings are AUTO-CONFIRMED and bills are generated immediately.

    FLOW:
    1. App submits reading → MeterReading created (is_confirmed=True)
    2. Bill is automatically generated with status='Pending'
    3. Admin processes payment → Bill status='Paid'

    Returns bill details:
    - status, message
    - consumer_name, account_number, reading_date
    - previous_reading, current_reading, consumption
    - rate, total_amount, field_staff_name
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        # Parse JSON data from the request body
        data = json.loads(request.body.decode('utf-8'))

        # Extract data from the request - MATCHING ANDROID APP FORMAT
        consumer_id = data.get('consumer_id') # Expecting consumer ID from app
        reading_value = data.get('reading')   # Expecting 'reading' key from app
        # Optional: Check if reading_date is sent, otherwise default to today
        reading_date_str = data.get('reading_date') # Expecting 'reading_date' key from app, can be None initially

        # Validate required fields (assuming reading_date is sent by the app now, or use today's date)
        if consumer_id is None or reading_value is None:
            return JsonResponse({'error': 'Missing required fields: consumer_id or reading'}, status=400)

        # Get the consumer based on ID (as sent by the app)
        try:
            consumer = Consumer.objects.get(id=consumer_id) # Use id instead of account_number
        except Consumer.DoesNotExist:
            return JsonResponse({'error': 'Consumer not found'}, status=404)

        # Check if consumer is disconnected - reject meter reading submission
        if consumer.status == 'disconnected':
            return JsonResponse({
                'error': 'Consumer is disconnected',
                'message': f'{consumer.first_name} {consumer.last_name} is currently disconnected. Meter reading not allowed.',
                'consumer_status': 'disconnected'
            }, status=403)

        # Determine the reading date
        if reading_date_str:
            # Parse the date string if provided by the app
            try:
                reading_date = timezone.datetime.strptime(reading_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)
        else:
            # Use the current date if no date is provided by the app
            reading_date = timezone.now().date()

        # Validate reading value (should be a positive number, handle potential float from app)
        try:
            # Convert to int, assuming the app sends an integer or a float that represents an integer
            # If the app sends a float representing a non-integer reading, this might need adjustment
            current_reading = int(reading_value) # Convert float to int
            if current_reading < 0:
                raise ValueError("Reading value cannot be negative")
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid reading value. Must be a non-negative number.'}, status=400)

        # Get previous reading
        previous_reading = get_previous_reading(consumer)

        # Calculate consumption
        consumption = current_reading - previous_reading

        # Validate consumption (current should be >= previous)
        if consumption < 0:
            return JsonResponse({
                'error': 'Invalid reading',
                'message': f'Current reading ({current_reading}) cannot be less than previous reading ({previous_reading})'
            }, status=400)

        # Calculate bill using tiered rates
        rate, total_amount, breakdown = calculate_water_bill(consumer, consumption)

        # Get field staff name
        field_staff_name = "System"  # Default
        if request.user.is_authenticated:
            field_staff_name = request.user.get_full_name() or request.user.username

        # --- NEW LOGIC: Check for existing unconfirmed reading on the same date ---
        try:
            existing_reading = MeterReading.objects.get(
                consumer=consumer,
                reading_date=reading_date
            )
            # If an existing reading is found for the same date
            if existing_reading.is_confirmed:
                # If it's already confirmed, don't allow updates
                error_msg = f"Reading for {consumer.account_number} on {reading_date} is already confirmed and cannot be updated via API."
                return JsonResponse({'error': error_msg}, status=400)
            else:
                # If it's unconfirmed, update the existing record
                existing_reading.reading_value = current_reading
                existing_reading.source = 'mobile_app' # Update source to reflect API submission
                existing_reading.save()

        except MeterReading.DoesNotExist:
            # If no existing reading for the date, create a new one
            # AUTO-CONFIRM: Reading is confirmed immediately and bill is generated
            reading = MeterReading.objects.create(
                consumer=consumer,
                reading_date=reading_date,
                reading_value=current_reading,
                source='mobile_app',  # Mark source as coming from the mobile app
                is_confirmed=True  # Auto-confirm - no manual confirmation needed
            )

            # ================================================================
            # AUTO-GENERATE BILL
            # ================================================================
            # Get previous confirmed reading for bill reference
            prev_reading_obj = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=reading_date
            ).order_by('-reading_date').first()

            # Get system settings for billing schedule
            setting = SystemSetting.objects.first()
            if setting:
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                billing_day = 1
                due_day = 20

            # Create Bill automatically with tiered rate info
            # Note: rate_per_cubic stores average rate, fixed_charge is 0 for tiered billing
            # (minimum charge is included in tier 1)
            Bill.objects.create(
                consumer=consumer,
                previous_reading=prev_reading_obj,
                current_reading=reading,
                billing_period=reading_date.replace(day=billing_day),
                due_date=reading_date.replace(day=due_day),
                consumption=consumption,
                rate_per_cubic=Decimal(str(rate)),  # Average rate for display
                fixed_charge=Decimal('0.00'),  # No separate fixed charge in tiered billing
                total_amount=Decimal(str(total_amount)),
                status='Pending'
            )

            # Create notification for new meter reading
            from .models import Notification
            from django.urls import reverse
            Notification.objects.create(
                user=None,  # Notify all admins
                notification_type='meter_reading',
                title='New Meter Reading Submitted',
                message=f'{consumer.first_name} {consumer.last_name} ({consumer.account_number}) - {consumer.barangay.name} | Bill: ₱{total_amount:.2f}',
                related_object_id=reading.id,
                redirect_url=reverse('consumers:barangay_meter_readings', kwargs={'barangay_id': consumer.barangay.id})
            )

        # Track activity for login session
        if request.user.is_authenticated:
            try:
                # Find current login session
                current_session = UserLoginEvent.objects.filter(
                    user=request.user,
                    logout_timestamp__isnull=True,
                    status='success'
                ).order_by('-login_timestamp').first()

                # Log the meter reading activity
                UserActivity.objects.create(
                    user=request.user,
                    action='meter_reading_submitted',
                    description=f"Meter reading submitted for {consumer.first_name} {consumer.last_name} ({consumer.account_number}). Reading: {current_reading}, Consumption: {consumption} m³",
                    login_event=current_session
                )
            except Exception:
                pass  # Don't fail the reading submission if activity logging fails

        # Return complete bill details (ALL 11 REQUIRED FIELDS)
        return JsonResponse({
            'status': 'success',
            'message': 'Reading submitted successfully',
            'consumer_name': f"{consumer.first_name} {consumer.last_name}",
            'account_number': consumer.account_number,
            'reading_date': str(reading_date),
            'previous_reading': int(previous_reading),
            'current_reading': int(current_reading),
            'consumption': int(consumption),
            'rate': rate,
            'total_amount': total_amount,
            'field_staff_name': field_staff_name
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON in request body'}, status=400)
    except Exception as e:
        # Log unexpected errors (Railway will capture this in logs)
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error submitting reading: {e}", exc_info=True)
        return JsonResponse({'error': 'Internal server error'}, status=500)





@csrf_exempt
def api_login(request):
    """Enhanced API login for Android app with security tracking."""
    from .decorators import get_client_ip, get_user_agent

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')

        # Get security information
        ip_address = get_client_ip(request)
        user_agent = get_user_agent(request)

        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            # Get staff's assigned barangay
            try:
                profile = StaffProfile.objects.get(user=user)

                # Record successful mobile login event
                UserLoginEvent.objects.create(
                    user=user,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    login_method='mobile',
                    status='success',
                    session_key=request.session.session_key
                )

                return JsonResponse({
                    'status': 'success',
                    'token': request.session.session_key,
                    'barangay': profile.assigned_barangay.name,
                    'user': {
                        'username': user.username,
                        'full_name': user.get_full_name()
                    }
                })
            except StaffProfile.DoesNotExist:
                # Still record the login even if there's no profile
                UserLoginEvent.objects.create(
                    user=user,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    login_method='mobile',
                    status='success'
                )
                return JsonResponse({'error': 'No assigned barangay'}, status=403)
        else:
            # Record failed login attempt
            if user:
                UserLoginEvent.objects.create(
                    user=user,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    login_method='mobile',
                    status='failed'
                )
            return JsonResponse({'error': 'Invalid credentials'}, status=401)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON in request body'}, status=400)
    except Exception as e:
        # Log unexpected errors (Railway will capture this in logs)
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error during API login: {e}", exc_info=True)
        return JsonResponse({'error': 'Internal server error'}, status=500)


@csrf_exempt
def api_logout(request):
    """API logout for Android app with session tracking."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        # Get token from request body or Authorization header
        token = None
        if request.body:
            try:
                data = json.loads(request.body)
                token = data.get('token')
            except json.JSONDecodeError:
                pass

        if not token:
            auth_header = request.headers.get('Authorization', '')
            if auth_header.startswith('Bearer '):
                token = auth_header[7:]

        if token:
            # Find and update the session
            latest_session = UserLoginEvent.objects.filter(
                session_key=token,
                logout_timestamp__isnull=True
            ).first()

            if latest_session:
                latest_session.logout_timestamp = timezone.now()
                latest_session.save()

                # Also logout from Django session if authenticated
                if request.user.is_authenticated:
                    logout(request)

                return JsonResponse({
                    'status': 'success',
                    'message': 'Logged out successfully',
                    'logout_time': timezone.now().isoformat()
                })
            else:
                # Session not found, might already be logged out
                return JsonResponse({
                    'status': 'success',
                    'message': 'Session already logged out or not found',
                    'logout_time': timezone.now().isoformat()
                })
        else:
            return JsonResponse({
                'error': 'No token provided',
                'message': 'Please provide token in request body or Authorization header'
            }, status=400)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error during API logout: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@login_required
def api_create_reading(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        consumer_id = data.get('consumer_id')
        reading_value = data.get('reading_value')
        reading_date = data.get('reading_date')  # YYYY-MM-DD

        # Validate consumer belongs to staff's barangay
        profile = StaffProfile.objects.get(user=request.user)
        consumer = Consumer.objects.get(id=consumer_id, barangay=profile.assigned_barangay)

        MeterReading.objects.create(
            consumer=consumer,
            reading_value=reading_value,
            reading_date=reading_date,
            source='field_app'
        )
        return JsonResponse({'status': 'success'})
    except StaffProfile.DoesNotExist:
        return JsonResponse({'error': 'Staff profile not found'}, status=403)
    except Consumer.DoesNotExist:
        return JsonResponse({'error': 'Consumer not found or not in assigned barangay'}, status=404)
    except Exception as e:
        # Log error but don't expose details
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"API create reading error: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create reading'}, status=400)

# consumers/views.py (Update the api_consumers function)


# ... (other imports remain the same) ...

@login_required
def api_consumers(request):
    """
    Get consumers for the staff's assigned barangay.
    OPTIMIZED: Uses prefetch_related and annotations to avoid N+1 queries.
    """
    try:
        from django.db.models import Prefetch, Exists, OuterRef

        profile = StaffProfile.objects.select_related('assigned_barangay').get(user=request.user)

        # PERFORMANCE FIX: Prefetch latest reading and annotate bill counts
        # This reduces 300+ queries to just 3 queries for 100 consumers
        consumers = Consumer.objects.filter(
            barangay=profile.assigned_barangay
        ).select_related(
            'barangay', 'purok'
        ).prefetch_related(
            # Prefetch only the latest confirmed reading per consumer
            Prefetch(
                'meter_readings',
                queryset=MeterReading.objects.filter(
                    is_confirmed=True
                ).order_by('-reading_date', '-created_at'),
                to_attr='latest_readings_list'
            )
        ).annotate(
            # Annotate pending bills count (1 query for all consumers)
            pending_bills_count_db=Count(
                'bills',
                filter=Q(bills__status='Pending')
            ),
            # Annotate delinquent status (no separate query needed)
            has_overdue_db=Exists(
                Bill.objects.filter(
                    consumer=OuterRef('pk'),
                    status='Pending',
                    due_date__lt=timezone.now().date()
                )
            )
        )

        data = []
        for consumer in consumers:
            # Get latest reading from prefetched data
            latest_reading_value = 0
            if consumer.latest_readings_list:
                latest_reading_value = consumer.latest_readings_list[0].reading_value

            # Append consumer data using annotated fields
            data.append({
                'id': consumer.id,
                'account_number': consumer.account_number,
                'name': f"{consumer.first_name} {consumer.last_name}",
                'first_name': consumer.first_name,
                'last_name': consumer.last_name,
                'serial_number': consumer.serial_number,
                'household_number': consumer.household_number,
                'barangay': consumer.barangay.name if consumer.barangay else '',
                'purok': consumer.purok.name if consumer.purok else '',
                'address': f"{consumer.purok.name if consumer.purok else ''}, {consumer.barangay.name if consumer.barangay else ''}",
                'phone_number': consumer.phone_number,
                'status': consumer.status,  # 'active' or 'disconnected'
                'is_active': consumer.status == 'active',
                'usage_type': consumer.usage_type,  # 'Residential' or 'Commercial' - needed for accurate rate calculation
                'latest_confirmed_reading': latest_reading_value,
                'previous_reading': latest_reading_value,  # Alias for Android app compatibility
                # Delinquent status from annotated field (no extra query!)
                'is_delinquent': consumer.has_overdue_db,
                'pending_bills_count': consumer.pending_bills_count_db
            })

        return JsonResponse(data, safe=False)
    except StaffProfile.DoesNotExist:
        return JsonResponse({'error': 'No assigned barangay'}, status=403)


@login_required
def api_get_previous_reading(request, consumer_id):
    """
    API endpoint to get the previous confirmed reading for a specific consumer.

    This provides a dedicated endpoint for the Android app to fetch just the
    previous reading without needing to load all consumers.

    URL: /api/consumers/<consumer_id>/previous-reading/
    Method: GET

    Returns:
        - consumer_id: int
        - account_number: str
        - consumer_name: str
        - previous_reading: int (0 if no confirmed reading exists)
        - last_reading_date: str or null
    """
    try:
        consumer = Consumer.objects.get(id=consumer_id)

        # Get previous reading using the same logic as get_previous_reading()
        latest_reading = MeterReading.objects.filter(
            consumer=consumer,
            is_confirmed=True
        ).order_by('-reading_date', '-created_at').first()

        previous_reading_value = latest_reading.reading_value if latest_reading else 0
        last_reading_date = latest_reading.reading_date.isoformat() if latest_reading else None

        return JsonResponse({
            'consumer_id': consumer.id,
            'account_number': consumer.account_number,
            'consumer_name': f"{consumer.first_name} {consumer.last_name}",
            'usage_type': consumer.usage_type,  # 'Residential' or 'Commercial' - needed for accurate rate calculation
            'previous_reading': previous_reading_value,
            'last_reading_date': last_reading_date
        })
    except Consumer.DoesNotExist:
        return JsonResponse({'error': 'Consumer not found'}, status=404)


# ... (other views remain the same) ...


# ======================
# SYSTEM SETTINGS
# ======================

# consumers/views.py




# consumers/views.py

# ... (other imports remain the same) ...
# ... (other imports remain the same) ...

# ... (your existing functions like api_login, api_consumers, api_submit_reading, system_management, etc.) ...

# NEW: API View for fetching the current water rates (Residential & Commercial)
@login_required # Ensure the user (app) is authenticated
def api_get_current_rates(request):
    """
    API endpoint for the Android app to fetch all tiered water rates.

    Returns complete tiered rate structure for both Residential and Commercial.
    """
    try:
        # Get the first (or only) SystemSetting object (singleton pattern)
        setting = SystemSetting.objects.first()
        if not setting:
            # Handle the case where no SystemSetting exists
            return JsonResponse({'error': 'System settings not configured.'}, status=500)

        # Return all tiered rates as JSON
        return JsonResponse({
            'status': 'success',
            # Residential Tiered Rates
            'residential': {
                'minimum_charge': float(setting.residential_minimum_charge),  # Tier 1: 1-5 m³
                'tier2_rate': float(setting.residential_tier2_rate),  # 6-10 m³
                'tier3_rate': float(setting.residential_tier3_rate),  # 11-20 m³
                'tier4_rate': float(setting.residential_tier4_rate),  # 21-50 m³
                'tier5_rate': float(setting.residential_tier5_rate),  # 51+ m³
            },
            # Commercial Tiered Rates
            'commercial': {
                'minimum_charge': float(setting.commercial_minimum_charge),  # Tier 1: 1-5 m³
                'tier2_rate': float(setting.commercial_tier2_rate),  # 6-10 m³
                'tier3_rate': float(setting.commercial_tier3_rate),  # 11-20 m³
                'tier4_rate': float(setting.commercial_tier4_rate),  # 21-50 m³
                'tier5_rate': float(setting.commercial_tier5_rate),  # 51+ m³
            },
            # Tier brackets info for reference
            'tier_brackets': {
                'tier1': '1-5 m³ (minimum charge)',
                'tier2': '6-10 m³',
                'tier3': '11-20 m³',
                'tier4': '21-50 m³',
                'tier5': '51+ m³'
            },
            # Legacy rates (for backward compatibility)
            'residential_rate_per_cubic': float(setting.residential_rate_per_cubic),
            'commercial_rate_per_cubic': float(setting.commercial_rate_per_cubic),
            'updated_at': setting.updated_at.isoformat()
        })

    except Exception as e:
        # Log unexpected errors
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching rates: {e}", exc_info=True)
        return JsonResponse({'error': 'Internal server error'}, status=500)

# ... (rest of your views) ...# consumers/views.py

# ... (other imports remain the same) ...

@login_required
@system_settings_permission_required
def system_management(request):
    """
    Manage system-wide settings: water rates, reading schedule, billing schedule, and penalties.
    RESTRICTED: Superuser only - Admins cannot access.
    """
    # Get the first (or only) SystemSetting instance (assumes singleton pattern)
    setting, created = SystemSetting.objects.get_or_create(id=1)

    if request.method == "POST":
        try:
            # Get water rate values
            new_res_rate_str = request.POST.get("residential_rate_per_cubic")
            new_comm_rate_str = request.POST.get("commercial_rate_per_cubic")
            new_fixed_charge_str = request.POST.get("fixed_charge")

            # Get reading schedule values
            reading_start = request.POST.get("reading_start_day")
            reading_end = request.POST.get("reading_end_day")

            # Get billing schedule values
            billing_day = request.POST.get("billing_day_of_month")
            due_day = request.POST.get("due_day_of_month")

            # Get penalty settings
            penalty_enabled = request.POST.get("penalty_enabled") == "on"
            penalty_type = request.POST.get("penalty_type", "percentage")
            penalty_rate_str = request.POST.get("penalty_rate", "10")
            fixed_penalty_str = request.POST.get("fixed_penalty_amount", "50")
            grace_period_str = request.POST.get("penalty_grace_period_days", "0")
            max_penalty_str = request.POST.get("max_penalty_amount", "500")

            # Validate and convert to Decimal
            new_res_rate = Decimal(new_res_rate_str)
            new_comm_rate = Decimal(new_comm_rate_str)
            new_fixed_charge = Decimal(new_fixed_charge_str)

            # Validate and convert penalty values
            penalty_rate = Decimal(penalty_rate_str)
            fixed_penalty = Decimal(fixed_penalty_str)
            grace_period = int(grace_period_str)
            max_penalty = Decimal(max_penalty_str)

            # Validate and convert schedule days
            reading_start = int(reading_start)
            reading_end = int(reading_end)
            billing_day = int(billing_day)
            due_day = int(due_day)

            # Validate rates
            if new_res_rate <= 0 or new_comm_rate <= 0 or new_fixed_charge < 0:
                raise ValueError("Rates must be positive and fixed charge cannot be negative.")

            # Validate penalty settings
            if penalty_rate < 0 or penalty_rate > 100:
                raise ValueError("Penalty rate must be between 0 and 100 percent.")
            if fixed_penalty < 0:
                raise ValueError("Fixed penalty amount cannot be negative.")
            if grace_period < 0 or grace_period > 30:
                raise ValueError("Grace period must be between 0 and 30 days.")
            if max_penalty < 0:
                raise ValueError("Maximum penalty cannot be negative.")
            if penalty_type not in ['percentage', 'fixed']:
                raise ValueError("Invalid penalty type.")

            # Validate all days are within 1-28
            for day, name in [(reading_start, "Reading start"), (reading_end, "Reading end"),
                              (billing_day, "Billing day"), (due_day, "Due day")]:
                if day < 1 or day > 28:
                    raise ValueError(f"{name} must be between 1 and 28.")

            # Validate reading period logic
            if reading_start > reading_end:
                raise ValueError("Reading start day must be before or equal to reading end day.")

            # Update the setting object
            setting.residential_rate_per_cubic = new_res_rate
            setting.commercial_rate_per_cubic = new_comm_rate
            setting.fixed_charge = new_fixed_charge
            setting.reading_start_day = reading_start
            setting.reading_end_day = reading_end
            setting.billing_day_of_month = billing_day
            setting.due_day_of_month = due_day

            # Update penalty settings
            setting.penalty_enabled = penalty_enabled
            setting.penalty_type = penalty_type
            setting.penalty_rate = penalty_rate
            setting.fixed_penalty_amount = fixed_penalty
            setting.penalty_grace_period_days = grace_period
            setting.max_penalty_amount = max_penalty

            setting.save()

            # Log the activity
            if hasattr(request, 'login_event') and request.login_event:
                UserActivity.objects.create(
                    user=request.user,
                    action='system_settings_updated',
                    description=f"Updated system settings including penalty configuration",
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    login_event=request.login_event
                )

            messages.success(request, "System settings updated successfully!")
        except (InvalidOperation, ValueError, TypeError) as e:
            messages.error(request, f"Invalid input: {e}")
        except Exception as e:
            messages.error(request, f"Error updating settings: {e}")

        return redirect("consumers:system_management")

    # For GET requests, pass the setting object to the template
    context = {
        "setting": setting,
    }
    return render(request, "consumers/system_management.html", context)

# ... (your other views remain the same) ...
# NEW: View for field staff to see their assigned consumers
@login_required
def consumer_list_for_staff(request):
    """Display consumers for the logged-in staff member's assigned barangay."""
    try:
        profile = StaffProfile.objects.get(user=request.user)
        assigned_barangay = profile.assigned_barangay
        consumers = Consumer.objects.filter(barangay=assigned_barangay)
    except StaffProfile.DoesNotExist:
        messages.error(request, "User profile not found. Please contact an administrator.")
        return redirect('consumers:login') # Or another appropriate page

    # Get login time from session
    login_time_iso = request.session.get('login_time')
    login_time_str = None
    if login_time_iso:
        try:
            login_time_obj = timezone.datetime.fromisoformat(login_time_iso.replace('Z', '+00:00'))
            login_time_str = login_time_obj.strftime("%b %d, %Y %H:%M:%S")
        except ValueError:
            login_time_str = "Unknown"

    context = {
        'consumers': consumers,
        'assigned_barangay': assigned_barangay,
        'login_time': login_time_str,
        'user': request.user, # Pass user for username display
    }
    return render(request, 'consumers/consumer_list_for_staff.html', context) # Create this template

# Example logout view

def user_logout(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('consumers:login') # Redirect to login page



# ======================
# EXPORT DELINQUENT CONSUMERS
# ======================

@login_required
def export_delinquent_consumers(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Use billing_period (not billing_date)
    bills = Bill.objects.filter(status='Pending')
    if month and year:
        bills = bills.filter(billing_period__month=month, billing_period__year=year)

    consumers = Consumer.objects.filter(bills__in=bills).distinct()

    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="delinquent_consumers_{month or "all"}_{year or "all"}.csv"'},
    )

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Middle Name', 'Last Name', 'Phone', 'Barangay', 'Serial', 'Pending Bills'])

    for consumer in consumers:
        pending_bills = consumer.bills.filter(status='Pending')
        total_pending = sum(b.total_amount for b in pending_bills)  # Use total_amount
        writer.writerow([
            consumer.first_name,
            consumer.middle_name or "",
            consumer.last_name,
            consumer.phone_number,
            consumer.barangay.name if consumer.barangay else "",
            consumer.serial_number,
            total_pending
        ])

    return response


# ======================
# SMART METER WEBHOOK
# ======================

@csrf_exempt
def smart_meter_webhook(request):
    """
    Webhook endpoint for IoT smart meters to submit readings.
    Requires API key authentication via X-API-Key header.
    Set SMART_METER_API_KEY in .env file.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)

    # Authenticate using API key from header
    from decouple import config
    expected_api_key = config('SMART_METER_API_KEY', default='')
    provided_api_key = request.META.get('HTTP_X_API_KEY', '')

    if not expected_api_key:
        # API key not configured - reject all requests for security
        return JsonResponse({'error': 'Webhook not configured'}, status=503)

    if provided_api_key != expected_api_key:
        # Invalid or missing API key
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    # Process the webhook data
    try:
        data = json.loads(request.body)

        # Validate required fields
        required_fields = ['consumer_id', 'reading', 'date']
        if not all(field in data for field in required_fields):
            return JsonResponse({
                'error': 'Missing required fields',
                'required': required_fields
            }, status=400)

        # Get consumer
        consumer = get_object_or_404(Consumer, id=data['consumer_id'])

        # Validate reading value
        reading_value = int(data['reading'])
        if reading_value < 0:
            return JsonResponse({'error': 'Reading value cannot be negative'}, status=400)

        # Create meter reading
        MeterReading.objects.create(
            consumer=consumer,
            reading_value=reading_value,
            reading_date=data['date'],
            source='smart_meter'
        )
        return JsonResponse({'status': 'success', 'message': 'Reading recorded'})

    except Consumer.DoesNotExist:
        return JsonResponse({'error': 'Consumer not found'}, status=404)
    except (ValueError, TypeError) as e:
        return JsonResponse({'error': 'Invalid data format'}, status=400)
    except Exception as e:
        # Log error but don't expose details
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Smart meter webhook error: {e}", exc_info=True)
        return JsonResponse({'error': 'Processing failed'}, status=500)


# ======================
# AUTH VIEWS
# ======================

@rate_limit_login
def staff_login(request):
    """Enhanced staff login with security tracking and rate limiting."""
    from .decorators import get_client_ip, get_user_agent

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        # Get security information
        ip_address = get_client_ip(request)
        user_agent = get_user_agent(request)

        if user is not None and user.is_staff:
            # Successful login
            login(request, user)

            # Record login event
            UserLoginEvent.objects.create(
                user=user,
                ip_address=ip_address,
                user_agent=user_agent,
                login_method='web',
                status='success',
                session_key=request.session.session_key
            )

            messages.success(request, f"Welcome back, {user.get_full_name() or user.username}!")
            return redirect('consumers:home')
        else:
            # Failed login attempt
            if user:
                # User exists but not staff - record failed attempt
                UserLoginEvent.objects.create(
                    user=user,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    login_method='web',
                    status='failed'
                )
            messages.error(request, "Invalid credentials or not staff member.")

    return render(request, 'consumers/login.html')


@login_required
def staff_logout(request):
    """Enhanced logout with session tracking."""
    # Update the latest active session for this user
    try:
        latest_session = UserLoginEvent.objects.filter(
            user=request.user,
            session_key=request.session.session_key,
            logout_timestamp__isnull=True
        ).first()

        if latest_session:
            latest_session.logout_timestamp = timezone.now()
            latest_session.save()
    except Exception as e:
        # Log error but don't prevent logout
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error updating logout timestamp: {e}")

    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect("consumers:staff_login")


# ======================
# PROFILE MANAGEMENT
# ======================

@login_required
def edit_profile(request):
    """
    Allow admin users to edit their profile and upload photo.
    """
    from .decorators import get_client_ip, get_user_agent

    try:
        profile = request.user.staffprofile
    except StaffProfile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('consumers:home')

    # Only allow admin to edit profile
    if profile.role != 'admin':
        messages.error(request, "Only administrators can edit their profile.")
        return redirect('consumers:home')

    if request.method == 'POST':
        updated = False

        # Update user information
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()

        if first_name and first_name != request.user.first_name:
            request.user.first_name = first_name
            updated = True
        if last_name and last_name != request.user.last_name:
            request.user.last_name = last_name
            updated = True
        if email and email != request.user.email:
            request.user.email = email
            updated = True

        if updated:
            request.user.save()

        # Handle profile photo upload
        if 'profile_photo' in request.FILES:
            photo = request.FILES['profile_photo']

            # Delete old photo if exists
            if profile.profile_photo:
                profile.profile_photo.delete(save=False)

            # Save new photo
            profile.profile_photo = photo
            profile.save()

            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='user_updated',
                description=f'{request.user.username} updated profile photo',
                ip_address=get_client_ip(request),
                user_agent=get_user_agent(request),
                target_user=request.user
            )

            messages.success(request, "Profile photo updated successfully!")
            updated = True

        if updated:
            messages.success(request, "Profile updated successfully!")
        else:
            messages.info(request, "No changes were made.")

        return redirect('consumers:edit_profile')

    return render(request, 'consumers/edit_profile.html', {
        'profile': profile
    })


# ======================
# PASSWORD RECOVERY
# ======================

def forgot_password_request(request):
    """
    Password reset request page for superuser/admin accounts.
    Sends secure reset token via email to the user's registered Gmail account.
    """
    from .decorators import get_client_ip, get_user_agent
    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags

    # Check if email is configured before processing any requests
    if not settings.EMAIL_HOST_USER or not settings.EMAIL_HOST_PASSWORD:
        messages.error(request, "Password reset via email is currently unavailable. Please contact your system administrator for password assistance.")
        return render(request, 'consumers/forgot_password.html', {'email_disabled': True})

    if request.method == "POST":
        username = request.POST.get('username')

        try:
            user = User.objects.get(username=username)

            # Check if user has an email address
            if not user.email:
                messages.error(request, "No email address found for this account. Please contact your administrator.")
                return redirect('consumers:forgot_password')

            # Only allow password reset for superuser and admin staff
            if not (user.is_superuser or (user.is_staff and user.groups.filter(name='Admin').exists())):
                messages.error(request, "Password reset is only available for administrators. Please contact your administrator for password assistance.")
                return redirect('consumers:forgot_password')

            # Check if user already has a valid token
            existing_token = PasswordResetToken.objects.filter(
                user=user,
                is_used=False,
                expires_at__gt=timezone.now()
            ).first()

            if existing_token:
                # Use existing valid token
                token = existing_token
            else:
                # Create new password reset token
                token = PasswordResetToken.objects.create(
                    user=user,
                    ip_address=get_client_ip(request)
                )

            # Build reset URL
            reset_url = request.build_absolute_uri(
                reverse('consumers:password_reset_confirm', kwargs={'token': token.token})
            )

            # Prepare email context
            email_context = {
                'username': user.username,
                'reset_url': reset_url,
                'request_time': token.created_at.strftime('%B %d, %Y at %I:%M %p'),
                'expiration_time': token.expires_at.strftime('%B %d, %Y at %I:%M %p'),
                'ip_address': get_client_ip(request) or 'Unknown',
            }

            # Render email templates
            html_message = render_to_string('consumers/emails/password_reset_email.html', email_context)
            plain_message = render_to_string('consumers/emails/password_reset_email.txt', email_context)

            # Send email
            try:
                subject = 'Password Reset Request - Balilihan Waterworks'
                # Gmail requires from_email to match EMAIL_HOST_USER
                from_email = settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL
                to_email = user.email

                # Log email attempt (for debugging)
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Attempting to send password reset email to {to_email} from {from_email}")

                # Create email with both HTML and plain text versions
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=plain_message,
                    from_email=from_email,
                    to=[to_email]
                )
                email.attach_alternative(html_message, "text/html")

                # Send with explicit connection settings
                email.send(fail_silently=False)

                # Log the activity
                UserActivity.objects.create(
                    user=user,
                    action='password_reset_requested',
                    description=f'Password reset email sent to {user.email}',
                    ip_address=get_client_ip(request),
                    user_agent=get_user_agent(request)
                )

                messages.success(request, f"Password reset link has been sent to your email: {user.email[:3]}***@{user.email.split('@')[1]}")
                return redirect('consumers:forgot_password')

            except Exception as e:
                # Log the error with detailed information
                import logging
                logger = logging.getLogger(__name__)
                error_msg = str(e)
                logger.error(f"Error sending password reset email to {user.email}: {error_msg}", exc_info=True)

                # Check for common email configuration issues
                if not settings.EMAIL_HOST_USER:
                    logger.error("EMAIL_HOST_USER is not configured!")
                    messages.error(request, "Email service is not configured. Please contact your administrator to set up EMAIL_HOST_USER.")
                elif not settings.EMAIL_HOST_PASSWORD:
                    logger.error("EMAIL_HOST_PASSWORD is not configured!")
                    messages.error(request, "Email service is not configured. Please contact your administrator to set up EMAIL_HOST_PASSWORD.")
                elif "authentication" in error_msg.lower() or "535" in error_msg:
                    logger.error("Gmail authentication failed - check EMAIL_HOST_PASSWORD (must be App Password)")
                    messages.error(request, "Email authentication failed. Please ensure Gmail App Password is configured correctly.")
                elif "connection" in error_msg.lower() or "timeout" in error_msg.lower():
                    logger.error("Connection to Gmail SMTP failed")
                    messages.error(request, "Could not connect to email server. Please try again later.")
                else:
                    messages.error(request, f"Failed to send password reset email. Error: {error_msg[:100]}")

                return redirect('consumers:forgot_password')

        except User.DoesNotExist:
            # For security, don't reveal if username exists or not
            messages.success(request, "If an account with that username exists, a password reset link has been sent to the registered email.")
            return redirect('consumers:forgot_password')

    return render(request, 'consumers/forgot_password.html')


def forgot_username(request):
    """
    Username recovery page - allows users to recover their username via email or full name.
    """
    recovered_username = None
    recovery_method = None

    if request.method == "POST":
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()

        if email:
            # Try to find user by email
            users = User.objects.filter(email__iexact=email, is_staff=True)
            if users.exists():
                if users.count() == 1:
                    recovered_username = users.first().username
                    recovery_method = 'email'
                    messages.success(request, f"Username found for email: {email}")
                else:
                    # Multiple users with same email
                    usernames = [u.username for u in users]
                    recovered_username = ", ".join(usernames)
                    recovery_method = 'email'
                    messages.success(request, f"Multiple accounts found for this email.")
            else:
                messages.error(request, "No staff account found with that email address.")

        elif first_name and last_name:
            # Try to find user by full name
            users = User.objects.filter(
                first_name__iexact=first_name,
                last_name__iexact=last_name,
                is_staff=True
            )
            if users.exists():
                if users.count() == 1:
                    recovered_username = users.first().username
                    recovery_method = 'name'
                    messages.success(request, f"Username found for {first_name} {last_name}")
                else:
                    usernames = [u.username for u in users]
                    recovered_username = ", ".join(usernames)
                    recovery_method = 'name'
                    messages.success(request, f"Multiple accounts found with this name.")
            else:
                messages.error(request, "No staff account found with that name.")
        else:
            messages.error(request, "Please provide either an email or your full name.")

    return render(request, 'consumers/forgot_username.html', {
        'recovered_username': recovered_username,
        'recovery_method': recovery_method
    })


def account_recovery(request):
    """
    Unified account recovery - recovers username and generates password reset link.
    """
    from .decorators import get_client_ip

    recovery_result = None

    if request.method == "POST":
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()

        user = None

        # Try to find user by email first
        if email:
            user = User.objects.filter(email__iexact=email, is_staff=True).first()
            if not user:
                messages.error(request, "No account found with that email address.")

        # Try by name if email not provided or not found
        elif first_name and last_name:
            user = User.objects.filter(
                first_name__iexact=first_name,
                last_name__iexact=last_name,
                is_staff=True
            ).first()
            if not user:
                messages.error(request, "No account found with that name.")
        else:
            messages.error(request, "Please enter your email or full name.")

        if user:
            # Check if user is admin/superuser (can reset password)
            can_reset = user.is_superuser or (user.is_staff and user.groups.filter(name='Admin').exists())

            recovery_result = {
                'username': user.username,
            }

            if can_reset:
                # Generate password reset token
                existing_token = PasswordResetToken.objects.filter(
                    user=user,
                    is_used=False,
                    expires_at__gt=timezone.now()
                ).first()

                if existing_token:
                    token = existing_token
                else:
                    token = PasswordResetToken.objects.create(
                        user=user,
                        ip_address=get_client_ip(request)
                    )

                reset_url = request.build_absolute_uri(
                    reverse('consumers:password_reset_confirm', kwargs={'token': token.token})
                )
                recovery_result['reset_url'] = reset_url

                # Log activity
                UserActivity.objects.create(
                    user=user,
                    action='password_reset_requested',
                    description=f'Account recovery initiated for {user.username}',
                    ip_address=get_client_ip(request)
                )

            messages.success(request, "Account found successfully!")

    return render(request, 'consumers/account_recovery.html', {
        'recovery_result': recovery_result
    })


def password_reset_confirm(request, token):
    """
    Confirm password reset with token and set new password.
    """
    from .decorators import get_client_ip, get_user_agent

    try:
        reset_token = PasswordResetToken.objects.get(token=token)

        # Check if token is valid
        if not reset_token.is_valid():
            messages.error(request, "This password reset link has expired or has already been used.")
            return redirect('consumers:staff_login')

        if request.method == "POST":
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            # Validate passwords match
            if new_password != confirm_password:
                messages.error(request, "Passwords do not match.")
                return render(request, 'consumers/reset_password.html', {
                    'token': token,
                    'username': reset_token.user.username
                })

            # Validate password strength
            if len(new_password) < 8:
                messages.error(request, "Password must be at least 8 characters long.")
                return render(request, 'consumers/reset_password.html', {
                    'token': token,
                    'username': reset_token.user.username
                })

            # Set new password
            user = reset_token.user
            user.set_password(new_password)
            user.save()

            # Mark token as used
            reset_token.is_used = True
            reset_token.save()

            # Log the activity
            UserActivity.objects.create(
                user=user,
                action='password_reset_completed',
                description=f'Password reset completed for {user.username}',
                ip_address=get_client_ip(request),
                user_agent=get_user_agent(request),
                target_user=user
            )

            messages.success(request, "Your password has been reset successfully! You can now login with your new password.")
            return redirect('consumers:password_reset_complete')

        return render(request, 'consumers/reset_password.html', {
            'token': token,
            'username': reset_token.user.username
        })

    except PasswordResetToken.DoesNotExist:
        messages.error(request, "Invalid password reset link.")
        return redirect('consumers:staff_login')


def password_reset_complete(request):
    """
    Password reset success confirmation page.
    """
    return render(request, 'consumers/reset_complete.html')


# ======================
# DASHBOARD
# ======================


@login_required
def home(request):
    """Staff dashboard showing key metrics and delinquent bills."""
    from datetime import timedelta
    from .models import Notification

    # Auto-cleanup: Delete notifications older than 1 month
    one_month_ago = datetime.now() - timedelta(days=30)
    Notification.objects.filter(created_at__lt=one_month_ago).delete()

    current_month = datetime.now().month
    current_year = datetime.now().year
    today = datetime.now().date()

    # Get counts
    connected_count = Consumer.objects.filter(status='active').count()
    disconnected_count = Consumer.objects.filter(status='disconnected').count()

    # Delinquent count (consumers with pending bills older than today)
    delinquent_count = Consumer.objects.filter(
        bills__status='Pending',
        bills__billing_period__lt=datetime.now().date()
    ).distinct().count()

    # ==========================================
    # REVENUE CALCULATIONS
    # ==========================================
    # Today's Revenue
    today_revenue = Payment.objects.filter(
        payment_date=today
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # This Month's Revenue
    monthly_revenue = Payment.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # Total Revenue (All Time)
    total_revenue = Payment.objects.aggregate(
        total=Sum('amount_paid')
    )['total'] or Decimal('0.00')

    # Today's payment count
    today_payment_count = Payment.objects.filter(payment_date=today).count()

    # Handle report filter - support both month_year and separate month/year
    month_year = request.GET.get('month_year')
    if month_year:
        try:
            selected_year, selected_month = month_year.split('-')
            selected_year = int(selected_year)
            selected_month = int(selected_month)
        except:
            selected_month = current_month
            selected_year = current_year
    else:
        selected_month = int(request.GET.get('month', current_month))
        selected_year = int(request.GET.get('year', current_year))

    # Get delinquent bills for the selected month/year
    delinquent_bills = Bill.objects.filter(
        status='Pending',
        billing_period__month=selected_month,
        billing_period__year=selected_year
    ).select_related('consumer', 'consumer__barangay').order_by('-billing_period')

    # Calculate total delinquent amount
    total_delinquent_amount = delinquent_bills.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    # Chart Data: Monthly Revenue Trend (Last 6 months)
    end_date = datetime.now().date()
    start_date = end_date - relativedelta(months=5)

    monthly_payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('amount_paid')
    ).order_by('month')

    revenue_labels = []
    revenue_data = []
    revenue_list = []  # For template iteration
    for item in monthly_payments:
        label = item['month'].strftime('%b %Y')
        amount = float(item['total'] or 0)
        revenue_labels.append(label)
        revenue_data.append(amount)
        revenue_list.append((label, amount))

    # Chart Data: Payment Status Distribution
    total_bills = Bill.objects.filter(status__in=['Pending', 'Unpaid', 'Overdue']).count()  # Outstanding bills only
    paid_bills = Bill.objects.filter(status='Paid').count()
    pending_bills = Bill.objects.filter(status='Pending').count()

    # Get all barangays for filter dropdown
    all_barangays = Barangay.objects.all().order_by('name')

    # Consumer Bill Status Data - Get latest bill for each consumer
    from django.db.models import Subquery, OuterRef
    latest_bill_subquery = Bill.objects.filter(
        consumer=OuterRef('pk')
    ).order_by('-billing_period').values('id')[:1]

    consumers_with_bills = Consumer.objects.filter(
        status='active'
    ).annotate(
        latest_bill_id=Subquery(latest_bill_subquery)
    ).select_related('barangay')

    consumer_bill_status = []
    for consumer in consumers_with_bills:
        if consumer.latest_bill_id:
            try:
                latest_bill = Bill.objects.get(id=consumer.latest_bill_id)
                consumer_bill_status.append({
                    'account_number': consumer.account_number,
                    'consumer_name': consumer.full_name,
                    'barangay': consumer.barangay.name if consumer.barangay else 'N/A',
                    'barangay_id': consumer.barangay.id if consumer.barangay else None,
                    'latest_bill_date': latest_bill.billing_period,
                    'latest_bill_amount': float(latest_bill.total_amount),
                    'payment_status': latest_bill.status,
                })
            except Bill.DoesNotExist:
                pass
        else:
            consumer_bill_status.append({
                'account_number': consumer.account_number,
                'consumer_name': consumer.full_name,
                'barangay': consumer.barangay.name if consumer.barangay else 'N/A',
                'barangay_id': consumer.barangay.id if consumer.barangay else None,
                'latest_bill_date': None,
                'latest_bill_amount': 0,
                'payment_status': 'No Bill',
            })

    # Chart Data: Barangay Consumer Distribution
    barangay_data = Consumer.objects.values('barangay__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    barangay_labels = [item['barangay__name'] or 'Unassigned' for item in barangay_data]
    barangay_counts = [item['count'] for item in barangay_data]

    # Chart Data: Monthly Consumption Trend
    monthly_consumption = Bill.objects.filter(
        billing_period__gte=start_date
    ).annotate(
        month=TruncMonth('billing_period')
    ).values('month').annotate(
        total_consumption=Sum('consumption')
    ).order_by('month')

    consumption_labels = []
    consumption_data = []
    for item in monthly_consumption:
        consumption_labels.append(item['month'].strftime('%b %Y'))
        consumption_data.append(float(item['total_consumption'] or 0))

    # Create a date object for proper month/year formatting in template
    from datetime import date
    selected_date = date(selected_year, selected_month, 1)

    context = {
        'connected_count': connected_count,
        'disconnected_count': disconnected_count,
        'delinquent_count': delinquent_count,
        'delinquent_bills': delinquent_bills,
        'total_delinquent_amount': total_delinquent_amount,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'selected_date': selected_date,  # For template date formatting
        'current_date': datetime.now(),  # For dynamic date display
        # Revenue data
        'today_revenue': today_revenue,
        'monthly_revenue': monthly_revenue,
        'total_revenue': total_revenue,
        'today_payment_count': today_payment_count,
        # Chart data
        'revenue_labels': json.dumps(revenue_labels),
        'revenue_data': json.dumps(revenue_data),
        'revenue_list': revenue_list,  # For template iteration
        'paid_bills': paid_bills,
        'pending_bills': pending_bills,
        'barangay_labels': json.dumps(barangay_labels),
        'barangay_counts': json.dumps(barangay_counts),
        'consumption_labels': json.dumps(consumption_labels),
        'consumption_data': json.dumps(consumption_data),
        'total_bills': total_bills,
        # Consumer Bill Status by Barangay
        'all_barangays': all_barangays,
        'consumer_bill_status': json.dumps(consumer_bill_status, default=str),
    }
    return render(request, 'consumers/home.html', context)




# ======================
# CONSUMER STATUS FILTERS
# ======================

@login_required
def connected_consumers(request):
    # Optimize query with select_related
    consumers = Consumer.objects.filter(status='active').select_related('barangay', 'purok')
    return render(request, 'consumers/consumer_list_filtered.html', {
        'title': 'Connected Consumers',
        'consumers': consumers
    })


# 1. LIST VIEW: Show all disconnected consumers (no ID needed)
@login_required
def disconnected_consumers_list(request):
    # Optimize query with select_related
    consumers = Consumer.objects.filter(status='disconnected').select_related('barangay', 'purok')
    return render(request, 'consumers/consumer_list_filtered.html', {
        'title': 'Disconnected Consumers',
        'consumers': consumers
    })

# 2. ACTION VIEW: Disconnect a specific consumer (requires ID)
@login_required
@disconnect_permission_required
def disconnect_consumer(request, consumer_id):
    """RESTRICTED: Superuser only - Admins cannot disconnect consumers."""
    consumer = get_object_or_404(Consumer, id=consumer_id)
    if request.method == 'POST':
        consumer.status = 'disconnected'
        consumer.disconnect_reason = request.POST.get('reason', '')
        consumer.save()

        # Track activity
        try:
            current_session = UserLoginEvent.objects.filter(
                user=request.user,
                logout_timestamp__isnull=True,
                status='success'
            ).order_by('-login_timestamp').first()

            UserActivity.objects.create(
                user=request.user,
                action='consumer_disconnected',
                description=f"Disconnected consumer: {consumer.first_name} {consumer.last_name} ({consumer.account_number}). Reason: {consumer.disconnect_reason or 'Not specified'}",
                login_event=current_session
            )
        except Exception:
            pass

        messages.success(request, f"{consumer.full_name} has been disconnected.")
        return redirect('consumers:disconnected_consumers')
    return render(request, 'consumers/confirm_disconnect.html', {'consumer': consumer})

@login_required
@disconnect_permission_required
def reconnect_consumer(request, consumer_id):
    """RESTRICTED: Superuser only - Admins cannot reconnect consumers."""
    consumer = get_object_or_404(Consumer, id=consumer_id)
    if request.method == 'POST':
        consumer.status = 'active'
        consumer.disconnect_reason = ''  # Optional: clear reason
        consumer.save()

        # Track activity
        try:
            current_session = UserLoginEvent.objects.filter(
                user=request.user,
                logout_timestamp__isnull=True,
                status='success'
            ).order_by('-login_timestamp').first()

            UserActivity.objects.create(
                user=request.user,
                action='consumer_reconnected',
                description=f"Reconnected consumer: {consumer.first_name} {consumer.last_name} ({consumer.account_number})",
                login_event=current_session
            )
        except Exception:
            pass

        messages.success(request, f"{consumer.full_name} has been reconnected.")
        return redirect('consumers:consumer_detail', consumer.id)
    # Optional: handle GET with confirmation, but POST-only is simpler
    return redirect('consumers:consumer_detail', consumer.id)


@login_required
def delinquent_consumers(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    bills = Bill.objects.filter(status='Pending')
    if month and year:
        bills = bills.filter(billing_period__month=month, billing_period__year=year)

    # Optimize query with select_related
    consumers = Consumer.objects.filter(bills__in=bills).select_related('barangay', 'purok').distinct()

    return render(request, 'consumers/consumer_list_filtered.html', {
        'title': 'Delinquent Consumers',
        'consumers': consumers,
        'selected_month': month,
        'selected_year': year
    })


@login_required
def delinquent_report_printable(request):
    """Printable delinquent report with receipt-style header"""
    from calendar import month_name

    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        messages.error(request, 'Month and year are required')
        return redirect('consumers:home')

    # Get all pending bills for the specified month/year
    bills = Bill.objects.filter(
        status='Pending',
        billing_period__month=month,
        billing_period__year=year
    ).select_related(
        'consumer__barangay',
        'consumer__purok',
        'previous_reading',
        'current_reading'
    ).order_by('consumer__account_number')

    # Calculate totals
    total_amount = sum(bill.total_amount for bill in bills)
    total_consumers = bills.count()

    # Format month display
    month_display = f"{month_name[int(month)]} {year}"

    context = {
        'bills': bills,
        'total_amount': total_amount,
        'total_consumers': total_consumers,
        'month_display': month_display,
        'month': month,
        'year': year,
        'generated_date': timezone.now()
    }

    return render(request, 'consumers/delinquent_report_print.html', context)


# ======================
# CONSUMER MANAGEMENT
# ======================


@login_required
def consumer_management(request):
    """Display consumer list with filters and modal form"""
    search_query = request.GET.get('search', '').strip()
    barangay_filter = request.GET.get('barangay', '').strip()

    # Optimize query with select_related to avoid N+1 queries
    consumers = Consumer.objects.select_related('barangay', 'purok', 'meter_brand').all()
    if search_query:
        # Search by name, account number, or serial number
        consumers = consumers.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(account_number__icontains=search_query) |
            Q(serial_number__icontains=search_query)
        )
    if barangay_filter:
        consumers = consumers.filter(barangay__id=barangay_filter)

    # Always pass a fresh form for the modal
    form = ConsumerForm()

    # Pagination
    paginator = Paginator(consumers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'consumers': page_obj,
        'form': form,
        'search_query': search_query,
        'barangays': Barangay.objects.all(),
        'barangay_filter': barangay_filter,
    }
    return render(request, 'consumers/consumer_management.html', context)


@login_required
@consumer_edit_permission_required
def add_consumer(request):
    """
    Handle adding a new consumer via full page form.
    RESTRICTED: Superuser only - Admins cannot create consumers.
    Includes duplicate detection to prevent accidental double-submission.
    """
    if request.method == "POST":
        form = ConsumerForm(request.POST)
        if form.is_valid():
            # ============================================================
            # DUPLICATE DETECTION - Prevent accidental double submission
            # ============================================================
            # Check if a consumer with same name, birth date, and barangay already exists
            # This catches duplicates from double-clicking the submit button
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            birth_date = form.cleaned_data.get('birth_date')
            barangay = form.cleaned_data.get('barangay')

            # Check for duplicate within the last 2 minutes (likely double-click)
            from datetime import timedelta
            from django.utils import timezone
            two_minutes_ago = timezone.now() - timedelta(minutes=2)

            duplicate = Consumer.objects.filter(
                first_name=first_name,
                last_name=last_name,
                birth_date=birth_date,
                barangay=barangay,
                created_at__gte=two_minutes_ago
            ).first()

            if duplicate:
                messages.warning(
                    request,
                    f"Consumer '{duplicate.first_name} {duplicate.last_name}' was already added (Account #: {duplicate.account_number}). "
                    "This may be a duplicate submission."
                )
                return redirect('consumers:consumer_management')

            # No duplicate found, proceed with creation
            consumer = form.save()

            # Track activity
            try:
                current_session = UserLoginEvent.objects.filter(
                    user=request.user,
                    logout_timestamp__isnull=True,
                    status='success'
                ).order_by('-login_timestamp').first()

                UserActivity.objects.create(
                    user=request.user,
                    action='consumer_created',
                    description=f"Created new consumer: {consumer.first_name} {consumer.last_name} ({consumer.account_number}) in {consumer.barangay.name}",
                    login_event=current_session
                )
            except Exception:
                pass  # Don't fail consumer creation if activity logging fails

            messages.success(request, f"Consumer added successfully! Account Number: {consumer.account_number}")
            return redirect('consumers:consumer_management')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ConsumerForm()

    return render(request, 'consumers/add_consumer.html', {'form': form})


@login_required
@consumer_edit_permission_required
def edit_consumer(request, consumer_id):
    """
    Edit consumer information.
    RESTRICTED: Superuser only - Admins cannot edit consumers.
    """
    consumer = get_object_or_404(Consumer, id=consumer_id)
    if request.method == 'POST':
        form = ConsumerForm(request.POST, instance=consumer)
        if form.is_valid():
            form.save()

            # Track activity
            try:
                current_session = UserLoginEvent.objects.filter(
                    user=request.user,
                    logout_timestamp__isnull=True,
                    status='success'
                ).order_by('-login_timestamp').first()

                UserActivity.objects.create(
                    user=request.user,
                    action='consumer_updated',
                    description=f"Updated consumer: {consumer.first_name} {consumer.last_name} ({consumer.account_number})",
                    login_event=current_session
                )
            except Exception:
                pass

            messages.success(request, "Consumer updated successfully!")
            return redirect('consumers:consumer_management')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ConsumerForm(instance=consumer)

    return render(request, 'consumers/edit_consumer.html', {'form': form, 'consumer': consumer})


@login_required
def consumer_list(request):
    """
    Enhanced consumer list view with filtering and statistics.
    """
    from datetime import datetime
    from django.db.models import Count

    # Base queryset with optimized queries
    consumers = Consumer.objects.select_related('barangay', 'purok', 'meter_brand').all()

    # Get all barangays for filter dropdown
    barangays = Barangay.objects.all().order_by('name')

    # Apply filters
    query = request.GET.get('q')
    barangay_filter = request.GET.get('barangay')
    status_filter = request.GET.get('status')

    if query:
        consumers = consumers.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(account_number__icontains=query) |
            Q(phone_number__icontains=query)
        )

    if barangay_filter:
        consumers = consumers.filter(barangay_id=barangay_filter)

    if status_filter:
        consumers = consumers.filter(status=status_filter)

    # Calculate statistics
    total_consumers = Consumer.objects.count()
    connected_count = Consumer.objects.filter(status='active').count()
    disconnected_count = Consumer.objects.filter(status='disconnected').count()

    # Consumers registered this month
    current_month = datetime.now().month
    current_year = datetime.now().year
    this_month_count = Consumer.objects.filter(
        registration_date__month=current_month,
        registration_date__year=current_year
    ).count()

    return render(request, 'consumers/consumer_list.html', {
        'consumers': consumers,
        'barangays': barangays,
        'total_consumers': total_consumers,
        'connected_count': connected_count,
        'disconnected_count': disconnected_count,
        'this_month_count': this_month_count,
    })


def consumer_detail(request, consumer_id):
    consumer = get_object_or_404(Consumer, id=consumer_id)
    latest_bills = consumer.bills.filter(status='Pending').order_by('-billing_period')[:3]
    return render(request, 'consumers/consumer_detail.html', {
        'consumer': consumer,
        'latest_bills': latest_bills
    })
# consumers/views.py

# ... other imports ...

# ... other view functions ...

# ... other imports ...


@login_required
def reports(request):
    """
    Simplified reports view - displays reports inline on the page.
    No AJAX, no modals, just simple form submission and display.
    """
    from datetime import datetime
    from calendar import month_name

    # Get parameters from GET or POST
    report_type = request.GET.get('report_type') or request.POST.get('report_type', 'revenue')
    month_year = request.GET.get('month_year') or request.POST.get('month_year')

    # Default to current month/year
    now = datetime.now()
    if month_year:
        try:
            selected_year = int(month_year.split('-')[0])
            selected_month = int(month_year.split('-')[1])
        except:
            selected_year = now.year
            selected_month = now.month
    else:
        selected_year = now.year
        selected_month = now.month

    # Format month/year for display
    month_year_value = f"{selected_year}-{selected_month:02d}"
    month_display = f"{month_name[selected_month]} {selected_year}"

    # Initialize report data
    report_data = None
    report_title = ""
    total_amount = 0
    record_count = 0

    # Generate report based on type
    if report_type == 'revenue':
        report_title = f"Revenue Report - {month_display}"
        report_data = Payment.objects.filter(
            payment_date__year=selected_year,
            payment_date__month=selected_month
        ).select_related('bill__consumer').order_by('payment_date')

        total_amount = report_data.aggregate(total=Sum('amount_paid'))['total'] or 0
        record_count = report_data.count()

    elif report_type == 'delinquency':
        report_title = f"Delinquent Accounts Report - {month_display}"
        report_data = Bill.objects.filter(
            billing_period__year=selected_year,
            billing_period__month=selected_month,
            status__in=['Pending', 'Overdue']
        ).select_related('consumer', 'consumer__barangay').order_by('consumer__account_number')

        total_amount = report_data.aggregate(total=Sum('total_amount'))['total'] or 0
        record_count = report_data.count()

    elif report_type == 'summary':
        report_title = f"Payment Summary Report - {month_display}"
        report_data = Payment.objects.filter(
            payment_date__year=selected_year,
            payment_date__month=selected_month
        ).values('bill__consumer__account_number').annotate(
            bill__consumer__full_name=Concat(
                'bill__consumer__first_name',
                Value(' '),
                'bill__consumer__middle_name',
                Value(' '),
                'bill__consumer__last_name'
            ),
            total_paid=Sum('amount_paid'),
            payment_count=Count('id')
        ).order_by('bill__consumer__account_number')

        total_amount = report_data.aggregate(total=Sum('total_paid'))['total'] or 0
        record_count = report_data.count()

    context = {
        'report_type': report_type,
        'month_year_value': month_year_value,
        'month_display': month_display,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'report_title': report_title,
        'report_data': report_data,
        'total_amount': total_amount,
        'record_count': record_count,
    }

    return render(request, 'consumers/reports.html', context)


@login_required
def export_report_excel(request):
    """Export report as Excel (.xlsx) file with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    report_type = request.GET.get('report_type', 'revenue')
    month_year = request.GET.get('month_year')

    if not month_year:
        return HttpResponse("Month/Year parameter required", status=400)

    selected_year = int(month_year.split('-')[0])
    selected_month = int(month_year.split('-')[1])

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Month name for display
    from calendar import month_name
    month_display = month_name[selected_month]

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if report_type == 'revenue':
        # Revenue Report
        ws.title = f"Revenue {month_display} {selected_year}"

        # Title
        ws['A1'] = "BALILIHAN WATERWORKS - REVENUE REPORT"
        ws['A1'].font = title_font
        ws['A2'] = f"Period: {month_display} {selected_year}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        # Headers
        headers = ['OR Number', 'Consumer Name', 'Account Number', 'Payment Date', 'Amount Paid', 'Change Given', 'Total Received']
        ws.append([])  # Empty row
        ws.append(headers)

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        payments = Payment.objects.filter(
            payment_date__year=selected_year,
            payment_date__month=selected_month
        ).select_related('bill__consumer').order_by('payment_date')

        total_amount = 0
        total_change = 0
        total_received = 0

        for payment in payments:
            consumer = payment.bill.consumer
            ws.append([
                payment.or_number,
                consumer.full_name,
                consumer.account_number,
                payment.payment_date.strftime('%Y-%m-%d'),
                float(payment.amount_paid),
                float(payment.change),
                float(payment.received_amount)
            ])
            total_amount += payment.amount_paid
            total_change += payment.change
            total_received += payment.received_amount

        # Total row
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = total_font
        ws[f'E{total_row}'] = float(total_amount)
        ws[f'F{total_row}'] = float(total_change)
        ws[f'G{total_row}'] = float(total_received)

        for col in ['A', 'E', 'F', 'G']:
            ws[f'{col}{total_row}'].font = total_font
            ws[f'{col}{total_row}'].fill = total_fill
            ws[f'{col}{total_row}'].border = border

        # Format currency columns
        for row in range(6, ws.max_row + 1):
            for col in ['E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '₱#,##0.00'
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

        filename = f"Revenue_Report_{month_display}_{selected_year}.xlsx"

    elif report_type == 'delinquency':
        # Delinquency Report
        ws.title = f"Delinquency {month_display}"

        # Title
        ws['A1'] = "BALILIHAN WATERWORKS - DELINQUENT ACCOUNTS REPORT"
        ws['A1'].font = title_font
        ws['A2'] = f"Period: {month_display} {selected_year}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        # Headers
        headers = ['Account Number', 'Consumer Name', 'Barangay', 'Billing Period', 'Due Date', 'Amount Due', 'Status']
        ws.append([])
        ws.append(headers)

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        bills = Bill.objects.filter(
            billing_period__year=selected_year,
            billing_period__month=selected_month,
            status__in=['Pending', 'Overdue']
        ).select_related('consumer__barangay').order_by('consumer__account_number')

        total_due = 0

        for bill in bills:
            consumer = bill.consumer
            ws.append([
                consumer.account_number,
                consumer.full_name,
                consumer.barangay.name if consumer.barangay else 'N/A',
                bill.billing_period.strftime('%B %Y'),
                bill.due_date.strftime('%Y-%m-%d'),
                float(bill.total_amount),
                bill.status
            ])
            total_due += bill.total_amount

        # Total row
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL DELINQUENT AMOUNT'
        ws[f'A{total_row}'].font = total_font
        ws[f'F{total_row}'] = float(total_due)
        ws[f'F{total_row}'].font = total_font
        ws[f'F{total_row}'].fill = total_fill
        ws[f'F{total_row}'].border = border
        ws[f'F{total_row}'].number_format = '₱#,##0.00'

        # Format currency column
        for row in range(6, ws.max_row):
            ws[f'F{row}'].number_format = '₱#,##0.00'
            ws[f'F{row}'].border = border
            ws[f'F{row}'].alignment = Alignment(horizontal='right')
            for col in ['A', 'B', 'C', 'D', 'E', 'G']:
                ws[f'{col}{row}'].border = border

        filename = f"Delinquency_Report_{month_display}_{selected_year}.xlsx"

    elif report_type == 'summary':
        # Summary Report
        ws.title = f"Summary {month_display}"

        # Title
        ws['A1'] = "BALILIHAN WATERWORKS - PAYMENT SUMMARY REPORT"
        ws['A1'].font = title_font
        ws['A2'] = f"Period: {month_display} {selected_year}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        # Headers
        headers = ['Account Number', 'Consumer Name', 'Total Amount Paid', 'Number of Payments']
        ws.append([])
        ws.append(headers)

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        summary_data = Payment.objects.filter(
            payment_date__year=selected_year,
            payment_date__month=selected_month
        ).values('bill__consumer__account_number').annotate(
            bill__consumer__full_name=Concat(
                'bill__consumer__first_name',
                Value(' '),
                'bill__consumer__middle_name',
                Value(' '),
                'bill__consumer__last_name'
            ),
            total_paid=Sum('amount_paid'),
            count=Count('id')
        ).order_by('bill__consumer__account_number')

        total_amount = 0
        total_count = 0

        for item in summary_data:
            ws.append([
                item['bill__consumer__account_number'],
                item['bill__consumer__full_name'],
                float(item['total_paid']),
                item['count']
            ])
            total_amount += item['total_paid']
            total_count += item['count']

        # Total row
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = total_font
        ws[f'C{total_row}'] = float(total_amount)
        ws[f'D{total_row}'] = total_count

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{total_row}'].font = total_font
            ws[f'{col}{total_row}'].fill = total_fill
            ws[f'{col}{total_row}'].border = border

        # Format currency column
        for row in range(6, ws.max_row + 1):
            ws[f'C{row}'].number_format = '₱#,##0.00'
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border

        filename = f"Payment_Summary_{month_display}_{selected_year}.xlsx"

    else:
        return HttpResponse("Invalid report type", status=400)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ... other view functions ...
# ======================
# AJAX
# ======================

@login_required
def load_puroks(request):
    barangay_id = request.GET.get('barangay_id')
    puroks = Purok.objects.filter(barangay_id=barangay_id).order_by('name')
    purok_list = [{'id': p.id, 'name': p.name} for p in puroks]
    return JsonResponse(purok_list, safe=False)


# ======================
# METER READINGS (CORRECTED)
# ======================






def get_consumer_display_id(consumer):
    """Returns the consumer's account number as the display ID"""
    return consumer.account_number or f"bw-{consumer.id:05d}"



def meter_reading_overview(request):
    """
    Meter Reading Overview with Enhanced Statistics and Progress Tracking.

    Features:
    - Barangay-level summary with progress bars
    - Overall completion percentage
    - Excel export for summary report
    - Updated vs pending consumer counts
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse

    today = date.today()
    current_month = today.replace(day=1)
    export_excel = request.GET.get('export', '')

    # Get all barangays, annotate total consumers, and ORDER BY name for alphabetical sorting
    barangays = Barangay.objects.annotate(
        total_consumers=Count('consumer', distinct=True)
    ).order_by('name')

    barangay_data = []
    for b in barangays:
        # Only count active (connected) consumers
        consumer_ids = list(Consumer.objects.filter(barangay=b, status='active').values_list('id', flat=True))
        total_consumers_count = len(consumer_ids)

        if not consumer_ids:
            ready = not_updated = updated = 0
            completion_percentage = 0
        else:
            # Count ALL unconfirmed readings (not just current month)
            # This matches what barangay_meter_readings shows
            ready = MeterReading.objects.filter(
                consumer_id__in=consumer_ids,
                is_confirmed=False
            ).values('consumer_id').distinct().count()

            # Count consumers who have at least one reading this month
            updated_consumers = MeterReading.objects.filter(
                consumer_id__in=consumer_ids,
                reading_date__gte=current_month
            ).values_list('consumer_id', flat=True).distinct()
            updated = len(set(updated_consumers))
            not_updated = total_consumers_count - updated

            # Calculate completion percentage (based on updated readings)
            completion_percentage = (updated / total_consumers_count * 100) if total_consumers_count > 0 else 0

        barangay_data.append({
            'barangay': b,
            'ready_to_confirm': ready,
            'not_yet_updated': not_updated,
            'updated_count': updated,
            'total_consumers': total_consumers_count,
            'completion_percentage': round(completion_percentage, 1),
        })

    # Calculate summary statistics for the overview page
    total_barangays = len(barangay_data)
    total_consumers_sum = sum(item['total_consumers'] for item in barangay_data)
    total_ready_sum = sum(item['ready_to_confirm'] for item in barangay_data)
    total_pending_sum = sum(item['not_yet_updated'] for item in barangay_data)
    total_updated_sum = sum(item['updated_count'] for item in barangay_data)

    # Calculate overall completion percentage
    overall_completion_percentage = (total_updated_sum / total_consumers_sum * 100) if total_consumers_sum > 0 else 0

    # Excel Export
    if export_excel == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reading Overview"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Meter Reading Overview - {current_month.strftime('%B %Y')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Barangay', 'Total Consumers', 'Readings This Month', 'Not Updated', 'Progress %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for idx, item in enumerate(barangay_data, 4):
            ws.cell(row=idx, column=1, value=item['barangay'].name)
            ws.cell(row=idx, column=2, value=item['total_consumers'])
            ws.cell(row=idx, column=3, value=item['updated_count'])
            ws.cell(row=idx, column=4, value=item['not_yet_updated'])
            ws.cell(row=idx, column=5, value=f"{item['completion_percentage']}%")

        # Summary row
        summary_row = len(barangay_data) + 5
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_consumers_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=total_updated_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value=total_pending_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=f"{round(overall_completion_percentage, 1)}%").font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=meter_reading_overview_{today.strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    context = {
        'barangay_data': barangay_data,
        'current_month': current_month,
        'total_barangays': total_barangays,
        'total_consumers_sum': total_consumers_sum,
        'total_ready_sum': total_ready_sum,
        'total_pending_sum': total_pending_sum,
        'overall_completion_percentage': round(overall_completion_percentage, 1),
    }

    return render(request, 'consumers/meter_reading_overview.html', context)

# ───────────────────────────────────────
# NEW: Barangay-Specific Readings (Enhanced)
# Shows the latest reading per consumer in the barangay,
# regardless of reading date or confirmation status.
# ───────────────────────────────────────
# consumers/views.py

@login_required
def barangay_meter_readings(request, barangay_id):
    barangay = get_object_or_404(Barangay, id=barangay_id)
    today = date.today()

    # Get all active consumers in this barangay
    consumers = Consumer.objects.filter(barangay=barangay, status='active').select_related('barangay').order_by('id')

    readings_with_data = []
    for consumer in consumers:
        # Get latest reading for this consumer (unconfirmed takes priority for display)
        latest_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date', '-created_at').first()

        if latest_reading:
            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=latest_reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                consumption = latest_reading.reading_value - prev.reading_value
            else:
                # First reading - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                consumption = latest_reading.reading_value - baseline
        else:
            latest_reading = None
            prev = None
            consumption = None

        readings_with_data.append({
            'consumer': consumer,
            'reading': latest_reading,
            'prev_reading': prev,
            'consumption': consumption,
            'display_id': get_consumer_display_id(consumer),
        })

    # Calculate counts for summary statistics
    pending_count = sum(1 for item in readings_with_data if item['reading'] and not item['reading'].is_confirmed)
    confirmed_count = sum(1 for item in readings_with_data if item['reading'] and item['reading'].is_confirmed)
    no_reading_count = sum(1 for item in readings_with_data if not item['reading'])

    return render(request, 'consumers/barangay_meter_readings.html', {
        'barangay': barangay,
        'readings': readings_with_data,
        'today': today,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'no_reading_count': no_reading_count,
    })


# ───────────────────────────────────────
# NEW: Confirm All Readings in Barangay
# ───────────────────────────────────────
@login_required
def confirm_all_readings(request, barangay_id):
    if request.method != "POST":
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    barangay = get_object_or_404(Barangay, id=barangay_id)
    readings_to_confirm = MeterReading.objects.filter(
        consumer__barangay=barangay,
        is_confirmed=False
    ).select_related('consumer')

    success_count = 0
    for reading in readings_to_confirm:
        try:
            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=reading.consumer,
                is_confirmed=True,
                reading_date__lt=reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                # Has previous reading - validate current >= previous
                if reading.reading_value < prev.reading_value:
                    continue
                cons = reading.reading_value - prev.reading_value
            else:
                # First reading for this consumer - use consumer's first_reading as baseline
                baseline = reading.consumer.first_reading if reading.consumer.first_reading else 0
                if reading.reading_value < baseline:
                    continue
                cons = reading.reading_value - baseline

            # Get system settings and determine rate based on usage type
            setting = SystemSetting.objects.first()
            if setting:
                # Use appropriate rate based on consumer's usage type (case-insensitive)
                if reading.consumer.usage_type and reading.consumer.usage_type.lower() == 'commercial':
                    rate = setting.commercial_rate_per_cubic
                else:  # residential or default
                    rate = setting.residential_rate_per_cubic
                fixed = setting.fixed_charge
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                # Fallback to defaults if no settings exist
                rate = Decimal('22.50')
                fixed = Decimal('50.00')
                billing_day = 1
                due_day = 20

            total = (Decimal(cons) * rate) + fixed

            Bill.objects.create(
                consumer=reading.consumer,
                previous_reading=prev,  # Will be None for first reading
                current_reading=reading,
                billing_period=reading.reading_date.replace(day=billing_day),
                due_date=reading.reading_date.replace(day=due_day),
                consumption=cons,
                rate_per_cubic=rate,
                fixed_charge=fixed,
                total_amount=total,
                status='Pending'
            )
            reading.is_confirmed = True
            reading.save()
            success_count += 1
        except Exception as e:
            # Log the error for debugging
            import logging
            logging.error(f"Error confirming reading {reading.id}: {str(e)}")
            continue

    if success_count > 0:
        messages.success(request, f"✅ {success_count} readings confirmed and bills generated.")
    return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)


@login_required
def confirm_all_readings_global(request):
    """
    Confirm ALL unconfirmed readings across ALL barangays.
    """
    if request.method != "POST":
        return redirect('consumers:meter_readings')

    # Get all unconfirmed readings
    readings_to_confirm = MeterReading.objects.filter(
        is_confirmed=False
    ).select_related('consumer')

    success_count = 0
    error_count = 0

    for reading in readings_to_confirm:
        try:
            consumer = reading.consumer

            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                if reading.reading_value < prev.reading_value:
                    continue
                cons = reading.reading_value - prev.reading_value
            else:
                # First reading for this consumer - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                if reading.reading_value < baseline:
                    continue
                cons = reading.reading_value - baseline

            # Get system settings
            setting = SystemSetting.objects.first()
            if setting:
                if consumer.usage_type and consumer.usage_type.lower() == 'commercial':
                    rate = setting.commercial_rate_per_cubic
                else:
                    rate = setting.residential_rate_per_cubic
                fixed = setting.fixed_charge
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                rate = Decimal('22.50')
                fixed = Decimal('50.00')
                billing_day = 1
                due_day = 20

            total = (Decimal(cons) * rate) + fixed

            Bill.objects.create(
                consumer=consumer,
                previous_reading=prev,
                current_reading=reading,
                billing_period=reading.reading_date.replace(day=billing_day),
                due_date=reading.reading_date.replace(day=due_day),
                consumption=cons,
                rate_per_cubic=rate,
                fixed_charge=fixed,
                total_amount=total,
                status='Pending'
            )
            reading.is_confirmed = True
            reading.save()
            success_count += 1
        except Exception as e:
            import logging
            logging.error(f"Error confirming reading {reading.id}: {str(e)}")
            error_count += 1
            continue

    if success_count > 0:
        messages.success(request, f"✅ {success_count} readings confirmed and bills generated across all barangays.")
    else:
        messages.info(request, "No unconfirmed readings found.")

    if error_count > 0:
        messages.warning(request, f"⚠️ {error_count} readings could not be processed.")

    return redirect('consumers:meter_readings')


# ───────────────────────────────────────
# NEW: Export to Excel
# ───────────────────────────────────────
@login_required
def export_barangay_readings(request, barangay_id):
    barangay = get_object_or_404(Barangay, id=barangay_id)
    current_month = date.today().replace(day=1)

    # Get latest reading per consumer in this barangay
    consumer_ids = Consumer.objects.filter(barangay=barangay).values_list('id', flat=True)
    if not consumer_ids:
        readings = MeterReading.objects.none()
    else:
        latest_date_subq = MeterReading.objects.filter(
            consumer=OuterRef('pk')
        ).order_by().values('consumer').annotate(
            max_date=Max('reading_date')
        ).values('max_date')[:1]

        readings = MeterReading.objects.select_related('consumer').filter(
            consumer__barangay=barangay,
            reading_date=Subquery(latest_date_subq)
        ).order_by('consumer__id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{barangay.name} Readings"

    headers = [
        'ID Number',           # ← Changed from "Account ID"
        'Consumer Name',
        'Current',
        'Previous',
        'Consumption (m³)',
        'Date',
        'Status'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).fill = header_fill
        ws.cell(1, col).font = header_font

    for r in readings:
        prev = MeterReading.objects.filter(
            consumer=r.consumer,
            is_confirmed=True,
            reading_date__lt=r.reading_date
        ).order_by('-reading_date').first()
        cons = (r.reading_value - prev.reading_value) if prev else '—'
        display_id = get_consumer_display_id(r.consumer)

        ws.append([
            display_id,
            f"{r.consumer.first_name} {r.consumer.last_name}",
            r.reading_value,
            prev.reading_value if prev else '—',
            cons,
            r.reading_date.strftime('%Y-%m-%d'),
            'Confirmed' if r.is_confirmed else 'Pending'
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Readings_{barangay.name}_{current_month.strftime("%Y-%m")}.xlsx'
    wb.save(response)
    return response


# ───────────────────────────────────────
# YOUR ORIGINAL VIEWS (UNCHANGED)
# ───────────────────────────────────────

# ============================================================================
# CONFIRM READING VIEW - ADMIN MANUAL CONFIRMATION
# ============================================================================
# This is the CRITICAL step where admin manually confirms a meter reading.
# Upon confirmation:
# 1. Consumption is calculated (current - previous reading)
# 2. Bill is IMMEDIATELY generated with status='Pending'
# 3. Bill appears in Inquire/Payment page for payment processing
#
# TESTING FLOW:
# Step 1: Mobile app submits reading (is_confirmed=False)
# Step 2: Admin clicks "Confirm" button HERE → Bill created instantly
# Step 3: Admin goes to Inquire page → Pays the bill
#
# NOTE: billing_day_of_month and due_day_of_month from SystemSettings
# only set the DATES on the bill, they do NOT delay bill creation.
# Bill is created IMMEDIATELY when you click Confirm.
# ============================================================================
@login_required
def confirm_reading(request, reading_id):
    """
    Confirm a meter reading and generate a bill for the consumer.

    TESTING: Click this to instantly generate a bill from a pending reading.

    FLOW:
    1. App submits reading → MeterReading (is_confirmed=False)
    2. Admin clicks Confirm → THIS FUNCTION RUNS
    3. Bill created with status='Pending' → Appears in payment page
    4. Admin processes payment → Bill status='Paid'

    If this is the first reading for the consumer, consumption is calculated
    from consumer.first_reading (initial meter value at registration).
    """
    current = get_object_or_404(MeterReading, id=reading_id)
    consumer = current.consumer
    barangay_id = consumer.barangay.id

    # Check if the reading is already confirmed
    if current.is_confirmed:
        messages.error(request, "This reading is already confirmed and billed.")
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    # Validate date - cannot confirm future readings
    if current.reading_date > date.today():
        messages.error(request, "Reading date cannot be in the future.")
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    # Find the previous confirmed reading for consumption calculation
    previous = MeterReading.objects.filter(
        consumer=consumer,
        is_confirmed=True,
        reading_date__lt=current.reading_date
    ).order_by('-reading_date').first()

    # ========================================================================
    # CONSUMPTION CALCULATION
    # ========================================================================
    # consumption = current_reading - previous_reading
    # If no previous reading, use consumer.first_reading as baseline
    # ========================================================================
    if previous:
        # Use the previous confirmed reading for calculation
        if current.reading_value < previous.reading_value:
            messages.error(request, f"Current reading ({current.reading_value}) < previous ({previous.reading_value}).")
            return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)
        elif current.reading_value == previous.reading_value:
            messages.warning(request, "Zero consumption. Bill will be generated.")
        consumption = current.reading_value - previous.reading_value
    else:
        # FIRST READING for this consumer
        # Use consumer's initial first_reading (set at registration) as baseline
        baseline = consumer.first_reading if consumer.first_reading else 0
        if current.reading_value < baseline:
            messages.error(request, f"Current reading ({current.reading_value}) < initial reading ({baseline}).")
            return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)
        consumption = current.reading_value - baseline

    # ========================================================================
    # BILL GENERATION - HAPPENS IMMEDIATELY ON CONFIRM
    # ========================================================================
    # Bill is created RIGHT NOW, not on a schedule.
    # billing_day_of_month: Only affects billing_period date on the bill
    # due_day_of_month: Only affects due_date on the bill
    # These do NOT control WHEN bills are generated - that's instant on confirm.
    # ========================================================================
    try:
        setting = SystemSetting.objects.first()

        # Apply correct rate based on consumer usage type (Residential vs Commercial)
        if setting:
            if consumer.usage_type == 'Residential':
                rate = setting.residential_rate_per_cubic
            else:  # Commercial
                rate = setting.commercial_rate_per_cubic
            fixed_charge = setting.fixed_charge
            # These only set the DATE values on the bill, not when it's created
            billing_day = setting.billing_day_of_month
            due_day = setting.due_day_of_month
        else:
            # Fallback rates if SystemSetting doesn't exist
            rate = Decimal('22.50') if consumer.usage_type == 'Residential' else Decimal('25.00')
            fixed_charge = Decimal('50.00')
            billing_day = 1
            due_day = 20

        # Calculate total: (consumption × rate) + fixed_charge
        total_amount = (Decimal(consumption) * rate) + fixed_charge

        # CREATE BILL IMMEDIATELY - This is the key action
        Bill.objects.create(
            consumer=consumer,
            previous_reading=previous,  # Will be None if this is the first reading
            current_reading=current,
            # billing_period: First day of billing month (just for display/records)
            billing_period=current.reading_date.replace(day=billing_day),
            # due_date: When payment is due (just for display/records)
            due_date=current.reading_date.replace(day=due_day),
            consumption=consumption,
            rate_per_cubic=rate,
            fixed_charge=fixed_charge,
            total_amount=total_amount,
            status='Pending'
        )

        # Mark the reading as confirmed
        current.is_confirmed = True
        current.save()

        messages.success(request, f"✅ Bill successfully generated for {get_consumer_display_id(consumer)}!")

    except Exception as e:
        messages.error(request, f"Failed to generate bill: {str(e)}")
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

# consumers/views.py

@login_required
def confirm_selected_readings(request, barangay_id):
    if request.method != "POST":
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    barangay = get_object_or_404(Barangay, id=barangay_id)
    reading_ids = request.POST.getlist('reading_ids')
    consumer_ids = request.POST.getlist('consumer_ids')  # For consumers with no reading (optional)

    # Debug: Check what readings were selected
    if not reading_ids:
        messages.warning(request, "No readings were selected for confirmation.")
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    success_count = 0

    # Process selected readings
    error_count = 0
    error_details = []
    for reading_id in reading_ids:
        try:
            reading = MeterReading.objects.get(id=reading_id)
            if reading.is_confirmed:
                continue

            consumer = reading.consumer
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                if reading.reading_value < prev.reading_value:
                    continue
                cons = reading.reading_value - prev.reading_value
            else:
                # First reading - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                if reading.reading_value < baseline:
                    continue
                cons = reading.reading_value - baseline

            # Get system settings
            setting = SystemSetting.objects.first()
            if setting:
                # Use appropriate rate based on consumer's usage type (case-insensitive)
                if consumer.usage_type and consumer.usage_type.lower() == 'commercial':
                    rate = setting.commercial_rate_per_cubic
                else:
                    rate = setting.residential_rate_per_cubic
                fixed = setting.fixed_charge
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                rate = Decimal('22.50')
                fixed = Decimal('50.00')
                billing_day = 1
                due_day = 20

            total = (Decimal(cons) * rate) + fixed

            Bill.objects.create(
                consumer=consumer,
                previous_reading=prev,
                current_reading=reading,
                billing_period=reading.reading_date.replace(day=billing_day),
                due_date=reading.reading_date.replace(day=due_day),
                consumption=cons,
                rate_per_cubic=rate,
                fixed_charge=fixed,
                total_amount=total,
                status='Pending'
            )
            reading.is_confirmed = True
            reading.save()
            success_count += 1
        except Exception as e:
            import logging
            logging.error(f"Error confirming reading {reading_id}: {str(e)}")
            error_count += 1
            error_details.append(f"Reading {reading_id}: {str(e)}")
            continue

    if success_count > 0:
        messages.success(request, f"✅ {success_count} readings confirmed and bills generated.")
    elif error_count == 0 and success_count == 0:
        messages.info(request, "No new readings to confirm (may already be confirmed).")

    if error_count > 0:
        messages.error(request, f"⚠️ {error_count} readings failed: {'; '.join(error_details[:3])}")

    return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

# ============================================================================
# INQUIRE / PAYMENT VIEW - BILL INQUIRY AND PAYMENT PROCESSING
# ============================================================================
# This is the FINAL step in the billing flow where payments are processed.
#
# COMPLETE FLOW (for testing):
# 1. Mobile app submits reading → MeterReading (is_confirmed=False)
# 2. Admin confirms reading → Bill created (status='Pending')
# 3. Admin comes HERE → Selects consumer → Sees pending bill
# 4. Admin enters received amount → Payment processed → Bill status='Paid'
# 5. Receipt generated with OR number
#
# HOW TO TEST:
# 1. Submit reading from app (or create manually)
# 2. Go to Meter Readings → Barangay → Click "Confirm"
# 3. Come to this Inquire page → Select barangay → Select consumer
# 4. You'll see the pending bill → Enter payment amount → Submit
# ============================================================================
@login_required
def inquire(request):
    """
    Bill Inquiry and Payment Processing page with Penalty Support.

    TESTING: After confirming a meter reading, come here to pay the bill.

    GET: Display bill inquiry form
        - Select Barangay → Purok → Consumer
        - Shows pending bills for selected consumer
        - Calculates and displays any applicable penalties

    POST: Process payment
        - Creates Payment record with auto-generated OR number
        - Includes penalty amount in payment if applicable
        - Updates Bill status to 'Paid'
        - Redirects to receipt page
    """
    from .utils import calculate_penalty, update_bill_penalty, get_payment_breakdown

    # Get system settings for penalty calculation
    system_settings = SystemSetting.objects.first()

    if request.method == "POST":
        # ====================================================================
        # PAYMENT PROCESSING (Final step in billing flow)
        # ====================================================================
        bill_id = request.POST.get('bill_id')
        received_amount = request.POST.get('received_amount')
        waive_penalty = request.POST.get('waive_penalty') == 'true'
        waive_reason = request.POST.get('waive_reason', '').strip()

        if not bill_id or not received_amount:
            messages.error(request, "Missing bill or payment amount.")
            return redirect('consumers:payment_counter')

        try:
            bill = get_object_or_404(Bill, id=bill_id)
            received_amount = Decimal(received_amount)

            # Update penalty calculation before processing payment
            update_bill_penalty(bill, system_settings, save=True)

            # Handle penalty waiver if requested (admin only)
            if waive_penalty and bill.penalty_amount > 0:
                if request.user.is_superuser or (hasattr(request.user, 'staffprofile') and request.user.staffprofile.role == 'admin'):
                    bill.penalty_waived = True
                    bill.penalty_waived_by = request.user
                    bill.penalty_waived_reason = waive_reason or "Waived at payment"
                    bill.penalty_waived_date = timezone.now()
                    bill.save()
                else:
                    messages.warning(request, "Only administrators can waive penalties.")

            # Calculate total amount due (bill + effective penalty)
            original_bill_amount = bill.total_amount
            effective_penalty = bill.effective_penalty
            total_amount_due = bill.total_amount_due

            # Validate payment amount
            if received_amount < total_amount_due:
                messages.error(request, f"Insufficient payment. Total amount due is ₱{total_amount_due:,.2f} (Bill: ₱{original_bill_amount:,.2f} + Penalty: ₱{effective_penalty:,.2f}).")
                return redirect(f"{request.path}?consumer={bill.consumer.id}")

            # CREATE PAYMENT RECORD with penalty tracking
            # OR number format: OR-YYYYMMDD-XXXXXX (auto-generated)
            payment = Payment.objects.create(
                bill=bill,
                original_bill_amount=original_bill_amount,
                penalty_amount=effective_penalty,
                penalty_waived=bill.penalty_waived,
                days_overdue_at_payment=bill.days_overdue,
                amount_paid=total_amount_due,
                received_amount=received_amount,
                change=received_amount - total_amount_due,
                payment_date=timezone.now(),
                or_number=f"OR-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}",
                processed_by=request.user,
            )

            # MARK BILL AS PAID - This completes the billing cycle
            bill.status = 'Paid'
            bill.save()

            # Log the activity
            if hasattr(request, 'login_event') and request.login_event:
                UserActivity.objects.create(
                    user=request.user,
                    action='payment_processed',
                    description=f"Processed payment OR#{payment.or_number} for {bill.consumer.full_name}. Amount: ₱{total_amount_due:,.2f}" +
                               (f" (includes ₱{effective_penalty:,.2f} penalty)" if effective_penalty > 0 else ""),
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    login_event=request.login_event
                )

            penalty_msg = f" (includes ₱{effective_penalty:,.2f} penalty)" if effective_penalty > 0 else ""
            messages.success(request, f"Payment processed successfully! OR: {payment.or_number}{penalty_msg}")
            return redirect('consumers:payment_receipt', payment_id=payment.id)

        except (ValueError, InvalidOperation):
            messages.error(request, "Invalid payment amount.")
            return redirect(request.get_full_path())
        except Exception as e:
            messages.error(request, f"Error processing payment: {e}")
            return redirect(request.get_full_path())

    # ===== GET REQUEST - Direct consumer selection (no location filter) =====
    selected_consumer_id = request.GET.get('consumer')

    # Load all active consumers directly for faster transaction
    consumers = Consumer.objects.filter(status='active').select_related('barangay', 'purok').order_by('last_name', 'first_name')

    # Build consumer bills dictionary and track consumers with any bills
    consumer_bills = {}
    consumers_with_bills = set()
    for c in consumers:
        bill = c.bills.filter(status='Pending').order_by('-billing_period').first()
        if bill:
            # Update penalty for display
            update_bill_penalty(bill, system_settings, save=True)
        consumer_bills[c.id] = bill
        # Check if consumer has ever had any bills
        if c.bills.exists():
            consumers_with_bills.add(c.id)

    selected_consumer = None
    latest_bill = None
    payment_breakdown = None

    if selected_consumer_id:
        selected_consumer = get_object_or_404(Consumer, id=selected_consumer_id)
        latest_bill = selected_consumer.bills.filter(status='Pending').order_by('-billing_period').first()

        if latest_bill:
            # Get detailed payment breakdown including penalty
            payment_breakdown = get_payment_breakdown(latest_bill, system_settings)

    # Count total pending bills
    total_pending_bills = Bill.objects.filter(status='Pending').count()

    # Check if user can waive penalties
    can_waive_penalty = request.user.is_superuser or (
        hasattr(request.user, 'staffprofile') and request.user.staffprofile.role == 'admin'
    )

    context = {
        'consumers': consumers,
        'consumer_bills': consumer_bills,
        'consumers_with_bills': consumers_with_bills,
        'selected_consumer': selected_consumer,
        'latest_bill': latest_bill,
        'payment_breakdown': payment_breakdown,
        'total_pending_bills': total_pending_bills,
        'system_settings': system_settings,
        'can_waive_penalty': can_waive_penalty,
    }
    return render(request, 'consumers/inquire.html', context)

@login_required
def payment_receipt(request, payment_id):
    """
    Display a printable official receipt for a payment.
    Ensures the payment exists and belongs to a valid bill/consumer.
    Includes tiered rate breakdown for transparent billing.
    """
    from .utils import calculate_tiered_water_bill

    payment = get_object_or_404(
        Payment.objects.select_related('bill__consumer', 'bill__previous_reading', 'bill__current_reading'),
        id=payment_id
    )

    # Calculate previous reading value: use stored value or compute from current - consumption
    if payment.bill.previous_reading:
        previous_reading_value = payment.bill.previous_reading.reading_value
    else:
        previous_reading_value = payment.bill.current_reading.reading_value - payment.bill.consumption

    # Calculate tiered breakdown for receipt display
    _, _, breakdown = calculate_tiered_water_bill(
        consumption=payment.bill.consumption,
        usage_type=payment.bill.consumer.usage_type
    )

    return render(request, 'consumers/receipt.html', {
        'payment': payment,
        'previous_reading_value': previous_reading_value,
        'breakdown': breakdown
    })


@login_required
def payment_history(request):
    """
    Payment History view showing all payments with penalty tracking.
    Allows filtering by date range, consumer, and penalty status.
    """
    from django.core.paginator import Paginator
    from django.db.models import Sum, Q

    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    penalty_filter = request.GET.get('penalty', '')  # 'with_penalty', 'waived', 'no_penalty'

    # Base query with related data
    payments = Payment.objects.select_related(
        'bill__consumer__barangay',
        'processed_by'
    ).order_by('-payment_date')

    # Apply filters
    if search_query:
        payments = payments.filter(
            Q(bill__consumer__first_name__icontains=search_query) |
            Q(bill__consumer__last_name__icontains=search_query) |
            Q(bill__consumer__account_number__icontains=search_query) |
            Q(or_number__icontains=search_query)
        )

    if date_from:
        payments = payments.filter(payment_date__date__gte=date_from)

    if date_to:
        payments = payments.filter(payment_date__date__lte=date_to)

    if penalty_filter == 'with_penalty':
        payments = payments.filter(penalty_amount__gt=0, penalty_waived=False)
    elif penalty_filter == 'waived':
        payments = payments.filter(penalty_waived=True)
    elif penalty_filter == 'no_penalty':
        payments = payments.filter(penalty_amount=0)

    # Calculate statistics
    total_stats = payments.aggregate(
        total_collected=Sum('amount_paid'),
        total_penalties=Sum('penalty_amount'),
        total_bills=Sum('original_bill_amount')
    )

    # Count by penalty status
    penalty_counts = {
        'with_penalty': payments.filter(penalty_amount__gt=0, penalty_waived=False).count(),
        'waived': payments.filter(penalty_waived=True).count(),
        'no_penalty': payments.filter(penalty_amount=0).count(),
    }

    # Pagination
    paginator = Paginator(payments, 25)  # 25 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'payments': page_obj,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'penalty_filter': penalty_filter,
        'total_stats': total_stats,
        'penalty_counts': penalty_counts,
        'total_count': payments.count(),
    }

    return render(request, 'consumers/payment_history.html', context)


@login_required
def user_login_history(request):
    """
    Enhanced login history with filtering, search, and analytics.
    Restricted to superusers and admins for security.
    """
    from .decorators import admin_or_superuser_required
    from django.db.models import Count, Q
    from datetime import timedelta
    from django.core.paginator import Paginator

    # Security check - only admins and superusers
    if not (request.user.is_superuser or (hasattr(request.user, 'staffprofile') and request.user.staffprofile.role == 'admin')):
        messages.error(request, "Access Denied: Administrative privileges required to view login history.")
        return render(request, 'consumers/403.html', status=403)

    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    method_filter = request.GET.get('method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Base query - prefetch activities for session tracking
    login_events = UserLoginEvent.objects.select_related('user').prefetch_related('activities').all()

    # Apply filters
    if search_query:
        login_events = login_events.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )

    if status_filter:
        login_events = login_events.filter(status=status_filter)

    if method_filter:
        login_events = login_events.filter(login_method=method_filter)

    if date_from:
        login_events = login_events.filter(login_timestamp__gte=date_from)

    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
        date_to_end = date_to_obj.replace(hour=23, minute=59, second=59)
        login_events = login_events.filter(login_timestamp__lte=date_to_end)

    # Order by most recent
    login_events = login_events.order_by('-login_timestamp')

    # Analytics
    total_logins = login_events.count()
    successful_logins = login_events.filter(status='success').count()
    failed_logins = login_events.filter(status='failed').count()
    active_sessions = login_events.filter(status='success', logout_timestamp__isnull=True).count()

    # Recent activity (last 24 hours)
    last_24_hours = timezone.now() - timedelta(hours=24)
    recent_logins = login_events.filter(login_timestamp__gte=last_24_hours).count()

    # Top users
    top_users = User.objects.annotate(
        login_count=Count('userloginevent')
    ).filter(login_count__gt=0).order_by('-login_count')[:5]

    # Pagination
    paginator = Paginator(login_events, 25)  # 25 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'login_events': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'date_from': date_from,
        'date_to': date_to,
        # Analytics
        'total_logins': total_logins,
        'successful_logins': successful_logins,
        'failed_logins': failed_logins,
        'active_sessions': active_sessions,
        'recent_logins': recent_logins,
        'top_users': top_users,
    }
    return render(request, 'consumers/user_login_history.html', context)


@login_required
def session_activities(request, session_id):
    """
    View detailed activities for a specific login session.
    Shows all meter readings and actions performed during the session.
    """
    # Security check - only admins and superusers
    if not (request.user.is_superuser or (hasattr(request.user, 'staffprofile') and request.user.staffprofile.role == 'admin')):
        messages.error(request, "Access Denied: Administrative privileges required.")
        return render(request, 'consumers/403.html', status=403)

    login_event = get_object_or_404(UserLoginEvent, id=session_id)
    activities = UserActivity.objects.filter(login_event=login_event).order_by('-created_at')

    # Calculate session stats
    total_readings = activities.filter(action='meter_reading_submitted').count()

    context = {
        'login_event': login_event,
        'activities': activities,
        'total_readings': total_readings,
    }
    return render(request, 'consumers/session_activities.html', context)


@login_required
def consumer_bill(request, consumer_id):
    """
    Display all bills for a specific consumer with summary statistics.
    Optimized with select_related to reduce database queries.
    """
    from django.db.models import Sum
    from datetime import datetime

    consumer = get_object_or_404(Consumer, id=consumer_id)
    bills = consumer.bills.select_related(
        'current_reading__consumer',
        'previous_reading__consumer'
    ).order_by('-billing_period')

    # Calculate summary statistics
    total_billed = bills.aggregate(total=Sum('total_amount'))['total'] or 0
    outstanding_balance = bills.filter(status='Pending').aggregate(total=Sum('total_amount'))['total'] or 0
    outstanding_balance += bills.filter(status='Overdue').aggregate(total=Sum('total_amount'))['total'] or 0

    return render(request, 'consumers/consumer_bill.html', {
        'consumer': consumer,
        'bills': bills,
        'total_billed': total_billed,
        'outstanding_balance': outstanding_balance,
        'today': datetime.now(),
    })


# ======================
# USER MANAGEMENT (SECURE)
# ======================

@login_required
def admin_verification(request):
    """
    Admin verification - requires password re-entry before accessing user management.
    Provides extra security layer for sensitive operations.
    """
    from .decorators import get_client_ip
    from django.contrib.auth import authenticate

    # Only superusers and admins can access this page
    if not (request.user.is_superuser or (hasattr(request.user, 'staffprofile') and request.user.staffprofile.role == 'admin')):
        messages.error(request, "Access Denied: Administrative privileges required.")
        return render(request, 'consumers/403.html', status=403)

    if request.method == 'POST':
        password = request.POST.get('password', '')
        destination = request.POST.get('destination', 'user_management')

        # Verify the password
        user = authenticate(username=request.user.username, password=password)

        if user is not None and user == request.user:
            # Password verified - store verification in session with timestamp
            request.session['admin_verified'] = True
            request.session['admin_verified_time'] = timezone.now().isoformat()

            # Log the verification
            UserLoginEvent.objects.create(
                user=request.user,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                login_method='web',
                status='success',
                session_key=request.session.session_key
            )

            messages.success(request, "Admin verification successful!")

            # Redirect to requested destination
            if destination == 'django_admin':
                return redirect('/admin/')
            else:
                return redirect('consumers:user_management')
        else:
            # Failed verification
            messages.error(request, "Incorrect password. Verification failed.")
            UserLoginEvent.objects.create(
                user=request.user,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                login_method='web',
                status='failed'
            )

    return render(request, 'consumers/admin_verification.html')


@login_required
@user_management_permission_required
def user_management(request):
    """
    Custom user management interface with enhanced security.
    RESTRICTED: Superuser only - Admins cannot manage users.
    """
    from django.db.models import Count, Q
    from django.core.paginator import Paginator

    # Check if admin verification is required and not expired
    admin_verified = request.session.get('admin_verified', False)
    admin_verified_time_str = request.session.get('admin_verified_time')

    # Check if verification has expired (15 minutes = 900 seconds)
    verification_expired = False
    if admin_verified and admin_verified_time_str:
        try:
            from datetime import timedelta
            verified_time = timezone.datetime.fromisoformat(admin_verified_time_str)
            if timezone.is_naive(verified_time):
                verified_time = timezone.make_aware(verified_time)
            time_since_verification = timezone.now() - verified_time
            if time_since_verification > timedelta(minutes=15):
                verification_expired = True
                # Clear expired verification
                request.session.pop('admin_verified', None)
                request.session.pop('admin_verified_time', None)
        except (ValueError, TypeError):
            verification_expired = True

    if not admin_verified or verification_expired:
        # Redirect to verification page
        if verification_expired:
            messages.warning(request, "Admin verification expired. Please verify again.")
        else:
            messages.warning(request, "Admin verification required to access User Management.")
        return redirect('consumers:admin_verification')

    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')

    # Base query
    users = User.objects.all()

    # Apply filters
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if role_filter:
        if role_filter == 'superuser':
            users = users.filter(is_superuser=True)
        elif role_filter == 'staff':
            users = users.filter(is_staff=True, is_superuser=False)
        elif role_filter == 'regular':
            users = users.filter(is_staff=False, is_superuser=False)

    if status_filter:
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)

    # Annotate with login count
    users = users.annotate(
        login_count=Count('userloginevent')
    ).select_related('staffprofile').order_by('-date_joined')

    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    staff_users = User.objects.filter(is_staff=True).count()
    superusers = User.objects.filter(is_superuser=True).count()

    # Available barangays for assignment
    barangays = Barangay.objects.all()

    context = {
        'users': page_obj,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_users': total_users,
        'active_users': active_users,
        'staff_users': staff_users,
        'superusers': superusers,
        'barangays': barangays,
    }
    return render(request, 'consumers/user_management.html', context)


@login_required
@user_management_permission_required
def create_user(request):
    """
    Create a new user with security validations.
    RESTRICTED: Superuser only - Admins cannot create users.
    """
    from .decorators import check_password_strength
    from django.contrib.auth.hashers import make_password

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        assigned_barangay_id = request.POST.get('assigned_barangay')
        role = request.POST.get('role', 'field_staff')

        # Validation
        if not username or not password:
            messages.error(request, "Username and password are required.")
            return redirect('consumers:user_management')

        if password != password_confirm:
            messages.error(request, "Passwords do not match.")
            return redirect('consumers:user_management')

        # Check password strength
        is_strong, msg = check_password_strength(password)
        if not is_strong:
            messages.error(request, f"Weak password: {msg}")
            return redirect('consumers:user_management')

        # Check if username exists
        if User.objects.filter(username=username).exists():
            messages.error(request, f"Username '{username}' already exists.")
            return redirect('consumers:user_management')

        # Only superusers can create other superusers
        if is_superuser and not request.user.is_superuser:
            messages.error(request, "Access Denied: Only superusers can create other superusers.")
            return redirect('consumers:user_management')

        try:
            # Create user
            user = User.objects.create(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                is_staff=is_staff,
                is_superuser=is_superuser,
                is_active=True
            )
            user.set_password(password)
            user.save()

            # Create staff profile for staff users
            if is_staff:
                # Field staff requires barangay, admin doesn't
                if role == 'field_staff' and not assigned_barangay_id:
                    messages.warning(request, f"User '{username}' created but field staff should have an assigned barangay.")

                barangay = None
                if assigned_barangay_id:
                    barangay = Barangay.objects.get(id=assigned_barangay_id)

                StaffProfile.objects.create(
                    user=user,
                    assigned_barangay=barangay,
                    role=role
                )

            messages.success(request, f"User '{username}' created successfully!")
            return redirect('consumers:user_management')

        except Exception as e:
            messages.error(request, f"Error creating user: {str(e)}")
            return redirect('consumers:user_management')

    return redirect('consumers:user_management')


@login_required
@user_management_permission_required
def edit_user(request, user_id):
    """
    Edit user details with security checks.
    RESTRICTED: Superuser only - Admins cannot edit users.
    """
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '').strip()
        user.last_name = request.POST.get('last_name', '').strip()
        user.email = request.POST.get('email', '').strip()
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.is_active = request.POST.get('is_active') == 'on'

        # Only allow changing superuser status if current user is superuser
        if request.user.is_superuser:
            user.is_superuser = request.POST.get('is_superuser') == 'on'

        user.save()

        # Update staff profile
        assigned_barangay_id = request.POST.get('assigned_barangay')
        role = request.POST.get('role', 'field_staff')

        if assigned_barangay_id:
            barangay = Barangay.objects.get(id=assigned_barangay_id)
            profile, created = StaffProfile.objects.get_or_create(user=user)
            profile.assigned_barangay = barangay
            profile.role = role
            profile.save()

        messages.success(request, f"User '{user.username}' updated successfully!")
        return redirect('consumers:user_management')

    return redirect('consumers:user_management')


@login_required
@user_management_permission_required
def delete_user(request, user_id):
    """
    Delete a user with confirmation.
    RESTRICTED: Superuser only - Admins cannot delete users.
    """
    user = get_object_or_404(User, id=user_id)

    # Prevent self-deletion
    if user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('consumers:user_management')

    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f"User '{username}' deleted successfully!")
        return redirect('consumers:user_management')

    return redirect('consumers:user_management')


@login_required
def reset_user_password(request, user_id):
    """Reset user password (superuser and admin)."""
    from .decorators import check_password_strength

    if not (request.user.is_superuser or (hasattr(request.user, 'staffprofile') and request.user.staffprofile.role == 'admin')):
        messages.error(request, "Access Denied: Administrative privileges required to reset passwords.")
        return redirect('consumers:user_management')

    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')

        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('consumers:user_management')

        # Check password strength
        is_strong, msg = check_password_strength(new_password)
        if not is_strong:
            messages.error(request, f"Weak password: {msg}")
            return redirect('consumers:user_management')

        user.set_password(new_password)
        user.save()
        messages.success(request, f"Password reset successfully for user '{user.username}'!")
        return redirect('consumers:user_management')

    return redirect('consumers:user_management')


# ============================================================================
# NOTIFICATION VIEWS - Handle real-time notifications
# ============================================================================
@login_required
def mark_notification_read(request, notification_id):
    """Mark a single notification as read (AJAX endpoint)."""
    if request.method == 'POST':
        try:
            from .models import Notification

            # Get the notification
            notification = get_object_or_404(Notification, id=notification_id)

            # Check if user has permission to mark this notification
            # Can mark if: user is the recipient, or it's a global notification (user=None)
            if notification.user is None or notification.user == request.user:
                notification.mark_as_read()
                return JsonResponse({'status': 'success', 'message': 'Notification marked as read'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def mark_all_notifications_read(request):
    """Mark all notifications as read for the current user (AJAX endpoint)."""
    if request.method == 'POST':
        try:
            from .models import Notification
            from django.db.models import Q

            # Mark all notifications for this user or global notifications as read
            notifications = Notification.objects.filter(
                is_read=False
            ).filter(
                Q(user=request.user) | Q(user__isnull=True)
            )

            count = notifications.count()
            for notification in notifications:
                notification.mark_as_read()

            return JsonResponse({
                'status': 'success',
                'message': f'{count} notification(s) marked as read'
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


# ===========================
# DATABASE DOCUMENTATION VIEW
# ===========================
@login_required
def database_documentation(request):
    """Display database schema, tables, and test data in a user-friendly UI."""

    # Get database statistics
    context = {
        # Table counts
        'total_consumers': Consumer.objects.count(),
        'total_barangays': Barangay.objects.count(),
        'total_puroks': Purok.objects.count(),
        'total_meter_brands': MeterBrand.objects.count(),
        'total_meter_readings': MeterReading.objects.count(),
        'total_bills': Bill.objects.count(),
        'total_payments': Payment.objects.count(),
        'total_users': User.objects.count(),

        # Sample data - Barangays
        'barangays': Barangay.objects.all().order_by('name')[:10],

        # Sample data - Puroks
        'puroks': Purok.objects.select_related('barangay').all()[:10],

        # Sample data - Meter Brands
        'meter_brands': MeterBrand.objects.all(),

        # Sample data - Consumers
        'consumers': Consumer.objects.select_related('barangay', 'purok', 'meter_brand').all()[:8],

        # Sample data - Meter Readings
        'meter_readings': MeterReading.objects.select_related('consumer').order_by('-reading_date')[:10],

        # Sample data - Bills
        'bills': Bill.objects.select_related('consumer', 'current_reading', 'previous_reading').order_by('-billing_period')[:10],

        # Sample data - Payments
        'payments': Payment.objects.select_related('bill', 'bill__consumer').order_by('-payment_date')[:10],

        # System Settings
        'system_settings': SystemSetting.objects.first(),
    }

    # Calculate sample billing amounts for display
    if context['system_settings']:
        settings = context['system_settings']
        # Residential example (15 m³)
        residential_consumption = 15
        residential_water_charge = settings.residential_rate_per_cubic * residential_consumption
        residential_total = residential_water_charge + settings.fixed_charge

        # Commercial example (30 m³)
        commercial_consumption = 30
        commercial_water_charge = settings.commercial_rate_per_cubic * commercial_consumption
        commercial_total = commercial_water_charge + settings.fixed_charge

        context.update({
            'residential_consumption': residential_consumption,
            'residential_water_charge': residential_water_charge,
            'residential_total': residential_total,
            'commercial_consumption': commercial_consumption,
            'commercial_water_charge': commercial_water_charge,
            'commercial_total': commercial_total,
        })

    # Database schema information
    context['database_tables'] = [
            {
                'name': 'Barangay',
                'model': 'consumers_barangay',
                'description': 'Stores barangay (village) information',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'name', 'type': 'VARCHAR(100)', 'constraints': 'UNIQUE, NOT NULL', 'description': 'Barangay name'},
                ]
            },
            {
                'name': 'Purok',
                'model': 'consumers_purok',
                'description': 'Stores purok (zone) information within barangays',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'name', 'type': 'VARCHAR(100)', 'constraints': 'NOT NULL', 'description': 'Purok name'},
                    {'name': 'barangay_id', 'type': 'INTEGER', 'constraints': 'FOREIGN KEY', 'description': 'Reference to Barangay'},
                ]
            },
            {
                'name': 'MeterBrand',
                'model': 'consumers_meterbrand',
                'description': 'Stores water meter brand information',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'name', 'type': 'VARCHAR(100)', 'constraints': 'UNIQUE, NOT NULL', 'description': 'Meter brand name'},
                ]
            },
            {
                'name': 'Consumer',
                'model': 'consumers_consumer',
                'description': 'Main consumer information table with personal, household, and meter details',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'account_number', 'type': 'VARCHAR(20)', 'constraints': 'UNIQUE, AUTO', 'description': 'Format: BW-XXXXX'},
                    {'name': 'first_name', 'type': 'VARCHAR(50)', 'constraints': 'NOT NULL', 'description': 'First name'},
                    {'name': 'middle_name', 'type': 'VARCHAR(50)', 'constraints': 'NULL', 'description': 'Middle name'},
                    {'name': 'last_name', 'type': 'VARCHAR(50)', 'constraints': 'NOT NULL', 'description': 'Last name'},
                    {'name': 'birth_date', 'type': 'DATE', 'constraints': 'NOT NULL', 'description': 'Date of birth'},
                    {'name': 'gender', 'type': 'VARCHAR(10)', 'constraints': 'NOT NULL', 'description': 'Male/Female/Other'},
                    {'name': 'phone_number', 'type': 'VARCHAR(15)', 'constraints': 'NOT NULL', 'description': 'Contact number'},
                    {'name': 'status', 'type': 'VARCHAR(20)', 'constraints': "DEFAULT 'active'", 'description': 'active/disconnected'},
                ]
            },
            {
                'name': 'MeterReading',
                'model': 'consumers_meterreading',
                'description': 'Stores meter reading records with confirmation status',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'consumer_id', 'type': 'INTEGER', 'constraints': 'FOREIGN KEY', 'description': 'Reference to Consumer'},
                    {'name': 'reading_date', 'type': 'DATE', 'constraints': 'NOT NULL', 'description': 'Date of reading'},
                    {'name': 'reading_value', 'type': 'INTEGER', 'constraints': 'NOT NULL', 'description': 'Cumulative meter value'},
                    {'name': 'is_confirmed', 'type': 'BOOLEAN', 'constraints': 'DEFAULT FALSE', 'description': 'Confirmation status'},
                ]
            },
            {
                'name': 'Bill',
                'model': 'consumers_bill',
                'description': 'Stores billing information with consumption and payment status',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'consumer_id', 'type': 'INTEGER', 'constraints': 'FOREIGN KEY', 'description': 'Reference to Consumer'},
                    {'name': 'billing_period', 'type': 'DATE', 'constraints': 'NOT NULL', 'description': 'First day of billing month'},
                    {'name': 'consumption', 'type': 'INTEGER', 'constraints': 'NOT NULL', 'description': 'Water consumption (m³)'},
                    {'name': 'total_amount', 'type': 'DECIMAL(10,2)', 'constraints': 'NOT NULL', 'description': 'Total bill amount'},
                    {'name': 'status', 'type': 'VARCHAR(20)', 'constraints': "DEFAULT 'Pending'", 'description': 'Pending/Paid/Overdue'},
                ]
            },
            {
                'name': 'Payment',
                'model': 'consumers_payment',
                'description': 'Records all payment transactions with OR numbers',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'bill_id', 'type': 'INTEGER', 'constraints': 'FOREIGN KEY', 'description': 'Reference to Bill'},
                    {'name': 'amount_paid', 'type': 'DECIMAL(10,2)', 'constraints': 'NOT NULL', 'description': 'Bill amount'},
                    {'name': 'or_number', 'type': 'VARCHAR(50)', 'constraints': 'UNIQUE, AUTO', 'description': 'Official Receipt number'},
                    {'name': 'payment_date', 'type': 'DATETIME', 'constraints': 'AUTO', 'description': 'Payment timestamp'},
                ]
            },
            {
                'name': 'SystemSetting',
                'model': 'consumers_systemsetting',
                'description': 'System-wide configuration settings for billing rates',
                'fields': [
                    {'name': 'id', 'type': 'INTEGER', 'constraints': 'PRIMARY KEY', 'description': 'Auto-increment ID'},
                    {'name': 'residential_rate_per_cubic', 'type': 'DECIMAL(10,2)', 'constraints': 'DEFAULT 22.50', 'description': 'Residential rate (₱/m³)'},
                    {'name': 'commercial_rate_per_cubic', 'type': 'DECIMAL(10,2)', 'constraints': 'DEFAULT 25.00', 'description': 'Commercial rate (₱/m³)'},
                    {'name': 'fixed_charge', 'type': 'DECIMAL(10,2)', 'constraints': 'DEFAULT 50.00', 'description': 'Fixed monthly charge'},
                ]
            },
        ]

    return render(request, 'consumers/database_documentation.html', context)


# ======================
# EMAIL TEST (Debug)
# ======================

@login_required
def test_email(request):
    """
    Test email configuration - only accessible by superusers.
    Provides detailed debugging information for SMTP issues.
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Superuser access required'}, status=403)

    import smtplib
    import logging
    from django.core.mail import send_mail, get_connection
    from django.http import JsonResponse

    logger = logging.getLogger(__name__)

    # Collect configuration info (without password)
    config_info = {
        'EMAIL_BACKEND': settings.EMAIL_BACKEND,
        'EMAIL_HOST': settings.EMAIL_HOST,
        'EMAIL_PORT': settings.EMAIL_PORT,
        'EMAIL_USE_TLS': settings.EMAIL_USE_TLS,
        'EMAIL_USE_SSL': getattr(settings, 'EMAIL_USE_SSL', False),
        'EMAIL_HOST_USER': settings.EMAIL_HOST_USER or '(not set)',
        'EMAIL_HOST_PASSWORD': '***SET***' if settings.EMAIL_HOST_PASSWORD else '(not set)',
        'DEFAULT_FROM_EMAIL': settings.DEFAULT_FROM_EMAIL,
        'EMAIL_TIMEOUT': getattr(settings, 'EMAIL_TIMEOUT', 'default'),
    }

    result = {
        'config': config_info,
        'tests': [],
        'success': False,
        'error': None
    }

    # Test 1: Check if credentials are set
    if not settings.EMAIL_HOST_USER:
        result['tests'].append({
            'name': 'Check EMAIL_HOST_USER',
            'status': 'FAIL',
            'message': 'EMAIL_HOST_USER is empty. Set it in Vercel environment variables.'
        })
        result['error'] = 'EMAIL_HOST_USER not configured'
        return JsonResponse(result)
    else:
        result['tests'].append({
            'name': 'Check EMAIL_HOST_USER',
            'status': 'PASS',
            'message': f'Set to: {settings.EMAIL_HOST_USER}'
        })

    if not settings.EMAIL_HOST_PASSWORD:
        result['tests'].append({
            'name': 'Check EMAIL_HOST_PASSWORD',
            'status': 'FAIL',
            'message': 'EMAIL_HOST_PASSWORD is empty. Set it in Vercel environment variables.'
        })
        result['error'] = 'EMAIL_HOST_PASSWORD not configured'
        return JsonResponse(result)
    else:
        result['tests'].append({
            'name': 'Check EMAIL_HOST_PASSWORD',
            'status': 'PASS',
            'message': f'Set (length: {len(settings.EMAIL_HOST_PASSWORD)} chars)'
        })

    # Test 2: Try SMTP connection
    try:
        result['tests'].append({
            'name': 'SMTP Connection Test',
            'status': 'TESTING',
            'message': f'Connecting to {settings.EMAIL_HOST}:{settings.EMAIL_PORT}...'
        })

        # Create SMTP connection
        if settings.EMAIL_USE_TLS:
            server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT, timeout=30)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(settings.EMAIL_HOST, settings.EMAIL_PORT, timeout=30)

        result['tests'][-1]['status'] = 'PASS'
        result['tests'][-1]['message'] = 'Connected to Gmail SMTP server'

        # Test 3: Try authentication
        result['tests'].append({
            'name': 'SMTP Authentication',
            'status': 'TESTING',
            'message': 'Authenticating...'
        })

        server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

        result['tests'][-1]['status'] = 'PASS'
        result['tests'][-1]['message'] = 'Authentication successful!'

        server.quit()

        # Test 4: Try sending a test email
        if request.GET.get('send') == 'true':
            to_email = request.GET.get('to', request.user.email)
            result['tests'].append({
                'name': 'Send Test Email',
                'status': 'TESTING',
                'message': f'Sending to {to_email}...'
            })

            try:
                send_mail(
                    subject='Test Email - Balilihan Waterworks',
                    message='This is a test email from Balilihan Waterworks. If you received this, email is working correctly!',
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[to_email],
                    fail_silently=False,
                )
                result['tests'][-1]['status'] = 'PASS'
                result['tests'][-1]['message'] = f'Email sent successfully to {to_email}'
                result['success'] = True
            except Exception as e:
                result['tests'][-1]['status'] = 'FAIL'
                result['tests'][-1]['message'] = f'Send failed: {str(e)}'
                result['error'] = str(e)
        else:
            result['success'] = True
            result['tests'].append({
                'name': 'Ready to Send',
                'status': 'INFO',
                'message': 'Add ?send=true&to=email@example.com to send a test email'
            })

    except smtplib.SMTPAuthenticationError as e:
        result['tests'][-1]['status'] = 'FAIL'
        result['tests'][-1]['message'] = f'Authentication failed: {str(e)}'
        result['error'] = 'Gmail authentication failed. Make sure you are using an App Password, not your regular Gmail password.'
        result['help'] = [
            '1. Go to https://myaccount.google.com/security',
            '2. Enable 2-Step Verification if not already enabled',
            '3. Go to App passwords (https://myaccount.google.com/apppasswords)',
            '4. Generate a new App Password for "Mail"',
            '5. Use that 16-character password (without spaces) as EMAIL_HOST_PASSWORD'
        ]
    except smtplib.SMTPConnectError as e:
        result['tests'][-1]['status'] = 'FAIL'
        result['tests'][-1]['message'] = f'Connection failed: {str(e)}'
        result['error'] = 'Could not connect to Gmail SMTP server'
    except Exception as e:
        if result['tests']:
            result['tests'][-1]['status'] = 'FAIL'
            result['tests'][-1]['message'] = f'Error: {str(e)}'
        result['error'] = str(e)
        logger.error(f"Email test error: {e}", exc_info=True)

    return JsonResponse(result, json_dumps_params={'indent': 2})
